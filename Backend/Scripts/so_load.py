import openpyxl
import requests
import time
import sys
import json
from openpyxl.styles import PatternFill
from datetime import datetime, timezone, timedelta
from collections import defaultdict

# ==========================================
# AUTH SETUP
# ==========================================
BASE_URL  = "https://cat5-devl.adaptive.qad.com/clouderp"
TOKEN_URL = f"{BASE_URL}/oauth/token"

AUTH_PARAMS = {
    "client_id":  "afb97fd221925b87f01489aeb0e02e81",
    "username":   "demo",
    "password":   "qad",
    "grant_type": "password"
}

# ── Endpoint templates ─────────────────────────────────────────────────────
HEADER_CREATE_URL = f"{BASE_URL}/api/erp/salesOrderHeaders?viewUri=urn:be:com.qad.sales.salesorder.ISalesOrderHeader"
INIT_LINE_URL     = f"{BASE_URL}/api/erp/salesOrderLinesGrid?initialize=true&domainCode={{domain}}&salesOrderNumber={{so}}"
FIELD_CHANGE_URL  = f"{BASE_URL}/api/erp/salesOrderLines/fieldChange?fieldName={{fieldName}}&dataOperation=CREATE"
SYNC_LINE_URL     = f"{BASE_URL}/api/erp/salesOrderLinesGrid"

RED_FILL     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]


# ==========================================
# HELPERS
# ==========================================

def get_new_token():
    try:
        response = requests.post(TOKEN_URL, params=AUTH_PARAMS)
        response.raise_for_status()
        token = response.json().get("access_token")
        if token:
            return token
        print("❌ Failed to find access_token in response!")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error generating token: {e}")
        sys.exit(1)


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT00:00:00.000Z")
    if isinstance(val, (int, float)):
        try:
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=int(val))).strftime("%Y-%m-%dT00:00:00.000Z")
        except Exception:
            pass
    str_val = str(val).strip()
    if not str_val or str_val.lower() == "none":
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(str_val, fmt).strftime("%Y-%m-%dT00:00:00.000Z")
        except ValueError:
            continue
    return str_val


def today_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT00:00:00.000Z")


def sv(row_data, key, default=""):
    val = row_data.get(key, default)
    return str(val).strip() if val is not None else default


def fv(row_data, key, default=0.0):
    val = row_data.get(key, default)
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def iv(row_data, key, default=0):
    val = row_data.get(key, default)
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def make_headers(token):
    return {
        "Content-Type":  "application/json",
        "Authorization": f"Bearer {token}"
    }


def debug_request(label, url, payload=None, resp=None, resp_json=None):
    print(f"\n{'='*60}")
    print(f"🔷 {label}")
    print(f"   URL: {url}")
    if payload:
        print(f"📤 PAYLOAD:\n{json.dumps(payload, indent=2, default=str)}")
    if resp is not None:
        print(f"   HTTP Status: {resp.status_code}")
    if resp_json is not None:
        print(f"🔍 RESPONSE:\n{json.dumps(resp_json, indent=2, default=str)}")
    print(f"{'='*60}")


def get_with_retry(url, token_ref, label="GET", max_retries=3):
    attempt = 0
    while True:
        try:
            resp = requests.get(url, headers=make_headers(token_ref[0]))
            if resp.status_code == 401:
                print("  🔄 Token expired — refreshing...")
                token_ref[0] = get_new_token()
                continue
            try:
                rj = resp.json()
            except Exception:
                rj = {}
            debug_request(label, url, resp=resp, resp_json=rj)
            return resp, rj
        except requests.exceptions.SSLError as e:
            attempt += 1
            if attempt >= max_retries:
                raise
            print(f"  ⚠️  SSL error on {label} (attempt {attempt}/{max_retries}), retrying in 2s... {e}")
            time.sleep(2)
        except requests.exceptions.ConnectionError as e:
            attempt += 1
            if attempt >= max_retries:
                raise
            print(f"  ⚠️  Connection error on {label} (attempt {attempt}/{max_retries}), retrying in 2s... {e}")
            time.sleep(2)


def post_with_retry(url, payload, token_ref, label="POST", max_retries=3):
    attempt = 0
    while True:
        try:
            resp = requests.post(url, json=payload, headers=make_headers(token_ref[0]))
            if resp.status_code == 401:
                print("  🔄 Token expired — refreshing...")
                token_ref[0] = get_new_token()
                continue
            try:
                rj = resp.json()
            except Exception:
                rj = {}
            debug_request(label, url, payload=payload, resp=resp, resp_json=rj)
            return resp, rj
        except requests.exceptions.SSLError as e:
            attempt += 1
            if attempt >= max_retries:
                raise
            print(f"  ⚠️  SSL error on {label} (attempt {attempt}/{max_retries}), retrying in 2s... {e}")
            time.sleep(2)
        except requests.exceptions.ConnectionError as e:
            attempt += 1
            if attempt >= max_retries:
                raise
            print(f"  ⚠️  Connection error on {label} (attempt {attempt}/{max_retries}), retrying in 2s... {e}")
            time.sleep(2)


def mark_row(ws, row_idx, status_col, error_col, status_val, error_val, fill=None):
    for cell in ws[row_idx]:
        cell.fill = PatternFill(fill_type=None)
    ws.cell(row=row_idx, column=status_col, value=status_val)
    ws.cell(row=row_idx, column=error_col,  value=error_val)
    if fill:
        for cell in ws[row_idx]:
            cell.fill = fill


def ensure_columns(ws, *col_names):
    header_row = [cell.value for cell in ws[1]]
    for name in col_names:
        if name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=name)
            header_row.append(name)
    return header_row


def strip_keys(d):
    return {k.strip() if k else k: v for k, v in d.items()}


# ==========================================
# PAYLOAD BUILDER
# ==========================================

def build_header_payload(row_data):
    domain   = sv(row_data, "Domain Code")
    so_num   = sv(row_data, "SO Number")
    order_dt = to_date(row_data.get("Order Date")) or today_iso()
    due_dt   = to_date(row_data.get("Due Date"))   or today_iso()
    uri      = f"urn:be:com.qad.sales.salesorder.ISalesOrderHeader:{domain}.{so_num}"
    sold_to = sv(row_data, "Sold To Customer Code") or sv(row_data, "Bill To Customer Code")
    bill_to = sv(row_data, "Bill To Customer Code")
    ship_to = sv(row_data, "Ship To Customer Code")
    ship_via = sv(row_data, "Ship Via")
    freight_list = sv(row_data, "Freight List")
    freight_terms = sv(row_data, "Freight Terms")
    currency_code = sv(row_data, "Currency Code")
    site_code = sv(row_data, "Site Code")
    customer_code = sv(row_data, "Sold To Customer Code")
    credit_terms = sv(row_data, "Credit Terms")
    daybook_set = sv(row_data, "Daybook Set")

    return {
        "salesOrderHeaders": [
            {
                "uri":                    uri,
                "salesOrderNumber":       so_num,
                "domainCode":             domain,
                "soldToCustomerCode":     sold_to,
                "billToCustomerCode":     bill_to,
                "shipToCustomerCode":     ship_to,
                "siteCode":               site_code,
                "currencyCode":           currency_code,
                "daybookSetCode":         daybook_set,
                "creditTermsCode":        credit_terms,
                "shipVia":                ship_via,
                "freightListCode":        freight_list,
                "freightTermsCode":       freight_terms,
                "languageCode":           "us",
                "orderDate":              order_dt,
                "dueDate":                due_dt,
                "exchangeRate":           1,
                "exchangeRate2":          1,
                "exchangeRateType":       "",
                "isConfirmed":            True,
                "isFixedPrice":           True,
                "isTaxable":              False,
                "isPartialOK":            True,
                "isPrimarySO":            True,
                "isForecastConsumed":     True,
                "isReprice":              True,
                "isDisplayTaxAmounts":    True,
                "isPrintSalesOrder":      True,
                "isPrintPackList":        True,
                "isPrintInvoiceHistory":  True,
                "isProcessPostTrailer":   True,
                "isUsingConsignmentInventory": True,
                "maximumAgingDays":       90,
                "dataOperation":          "C",
                "concurrencyHash":        "",
                "disallowedActions":      "",
                "disallowedActionsMessage": "",
                "supplementaryMessages":  [],
                "SOSalespersons":         [],
                "SOSaveOptions":          [],
                "soDraftMappings":        [],
            }
        ]
    }


# ==========================================
# LINE CREATION FLOW
# ==========================================

def create_line(domain, so_num, line_row_data, line_number, token_ref):
    """
    Full stateful SO line creation:
    1. GET  initialize
    2. POST fieldChange(itemCode)
    3. POST fieldChange(siteCode)
    4. POST fieldChange(quantityOrdered)
    5. POST fieldChange(listPrice)         — only if > 0
    6. POST fieldChange(discountFormatted) — only if > 0
    7. POST fieldChange(netPrice)          — only if > 0
    8. POST fieldChange(dueDate)
    9. POST sync (salesOrderLinesGrid)
    """

    item_code = sv(line_row_data, "Item Code")
    site_code = sv(line_row_data, "Site Code")
    qty       = fv(line_row_data, "Quantity Ordered")
    price     = fv(line_row_data, "List Price")
    discount  = fv(line_row_data, "Discount", 0.0)
    net       = fv(line_row_data, "Net Price", 0.0)
    due_dt    = to_date(line_row_data.get("Due Date")) or today_iso()

    # ── 1. Initialize blank line ───────────────────────────────────────────
    print(f"\n    ── INIT LINE {line_number} ──")
    init_url = INIT_LINE_URL.format(domain=domain, so=so_num)
    resp, resp_json = get_with_retry(init_url, token_ref, label=f"INIT line {line_number}")

    if resp.status_code != 200:
        return False, f"INIT failed: HTTP {resp.status_code}"

    lines = resp_json.get("data", {}).get("salesOrderLines", [])
    if not lines:
        return False, "INIT returned no line object"

    line = lines[0]
    line["salesOrderLine"] = line_number
    print(f"    ✅ Init OK — line slot: {line.get('salesOrderLine')}")

    # ── Helper: run a fieldChange and chain the response ──────────────────
    def field_change(field_name, value):
        nonlocal line
        line[field_name] = value
        url = FIELD_CHANGE_URL.format(fieldName=field_name)
        r, rj = post_with_retry(url, {"salesOrderLines": [line]}, token_ref,
                                 label=f"fieldChange({field_name})")
        if r.status_code != 200:
            return False, f"fieldChange({field_name}) HTTP {r.status_code}"
        updated = rj.get("data", {}).get("salesOrderLines", [])
        if not updated:
            return False, f"fieldChange({field_name}) returned no line"
        line = updated[0]
        return True, ""

    # ── 2–8. Field changes in order ────────────────────────────────────────
    steps = [
        ("itemCode",        item_code),
        ("siteCode",        site_code),
        ("quantityOrdered", qty),
    ]
    if price > 0:
        steps.append(("listPrice", price))
    if discount > 0:
        steps.append(("discountFormatted", discount))
    if net > 0:
        steps.append(("netPrice", net))
    steps.append(("dueDate", due_dt))

    for field_name, value in steps:
        print(f"    🔧 fieldChange: {field_name} = {value}")
        ok, err = field_change(field_name, value)
        if not ok:
            return False, err
        print(f"    ✅ {field_name} OK")
        time.sleep(0.1)

    # ── 9. Sync / commit line ─────────────────────────────────────────────
    print(f"\n    ── SYNC LINE {line_number} ──")
    resp, resp_json = post_with_retry(
        SYNC_LINE_URL,
        {"salesOrderLines": [line]},
        token_ref,
        label=f"SYNC line {line_number}"
    )
    if resp.status_code == 200 and resp_json.get("submitResult", {}).get("success"):
        print(f"    ✅ Line {line_number} committed!")
        return True, ""
    else:
        errors    = resp_json.get("submitResult", {}).get("errors", [])
        error_msg = "; ".join(
            [f"{e.get('message','')} | field: {e.get('fieldName','')}" for e in errors if e.get("message")]
        )
        if not error_msg:
            error_msg = resp_json.get("message", f"HTTP {resp.status_code}")
        return False, f"Sync failed: {error_msg}"


# ==========================================
# MAIN RUN
# ==========================================

def run(file_path):
    print(f"📄 Processing: {file_path}")
    token_ref = [get_new_token()]

    wb = openpyxl.load_workbook(file_path)

    if "Header" not in wb.sheetnames or "Lines" not in wb.sheetnames:
        print("❌ Workbook must have sheets named 'Header' and 'Lines'")
        sys.exit(1)

    ws_h = wb["Header"]
    ws_l = wb["Lines"]

    h_headers = ensure_columns(ws_h, "Status", "Error")
    l_headers = ensure_columns(ws_l, "Status", "Error")

    h_headers = [h.strip() if h else h for h in h_headers]
    l_headers = [h.strip() if h else h for h in l_headers]

    h_status_col = h_headers.index("Status") + 1
    h_error_col  = h_headers.index("Error")  + 1
    l_status_col = l_headers.index("Status") + 1
    l_error_col  = l_headers.index("Error")  + 1

    # ── Index header rows ──────────────────────────────────────────────────
    header_rows = {}
    for row_idx, row in enumerate(ws_h.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        row_data = strip_keys(dict(zip(h_headers, row_values)))
        so_num   = sv(row_data, "SO Number")
        if so_num:
            header_rows[so_num] = (row_idx, row_data)

    # ── Index line rows ────────────────────────────────────────────────────
    line_rows = defaultdict(list)
    for row_idx, row in enumerate(ws_l.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        row_data = strip_keys(dict(zip(l_headers, row_values)))
        so_num   = sv(row_data, "SO Number")
        if so_num:
            line_rows[so_num].append((row_idx, row_data))

    h_success = h_fail = l_success = l_fail = 0

    # ── Process each SO ────────────────────────────────────────────────────
    for so_num, (h_row_idx, h_row_data) in header_rows.items():

        print(f"\n{'#'*60}")
        print(f"📦 PROCESSING SO: {so_num}")
        print(f"{'#'*60}")

        h_status = str(h_row_data.get("Status", "")).strip().upper()
        domain   = sv(h_row_data, "Domain Code")

        # ── Step 1: Create header ──────────────────────────────────────────
        if h_status == "DONE":
            print(f"  ⏭️  Skipping header (already DONE): {so_num}")
        else:
            payload = build_header_payload(h_row_data)
            print(f"\n  ── CREATE HEADER: {so_num} ──")
            try:
                resp, resp_json = post_with_retry(
                    HEADER_CREATE_URL, payload, token_ref,
                    label=f"CREATE HEADER {so_num}"
                )

                if resp.status_code == 200 and resp_json.get("submitResult", {}).get("success"):
                    mark_row(ws_h, h_row_idx, h_status_col, h_error_col, "DONE", "")
                    h_success += 1
                    print(f"  ✅ Header created: {so_num}")
                else:
                    errors = resp_json.get("submitResult", {}).get("errors", [])
                    error_msg = "; ".join(
                        [f"{e.get('message','')} | field: {e.get('fieldName','')}" for e in errors if e.get("message")]
                    )
                    if not error_msg:
                        error_msg = resp_json.get("message", f"HTTP {resp.status_code}")
                    mark_row(ws_h, h_row_idx, h_status_col, h_error_col, "ERROR", error_msg, RED_FILL)
                    h_fail += 1
                    print(f"  ❌ Header failed: {so_num} — {error_msg}")
                    wb.save(file_path)
                    continue

            except Exception as e:
                mark_row(ws_h, h_row_idx, h_status_col, h_error_col, "ERROR", str(e), RED_FILL)
                h_fail += 1
                print(f"  ⚠️  Header exception: {so_num} — {e}")
                wb.save(file_path)
                continue

            time.sleep(0.3)

        # ── Step 2: Create lines ───────────────────────────────────────────
        so_lines = line_rows.get(so_num, [])
        if not so_lines:
            print(f"  ⚠️  No lines found for SO: {so_num}")
            continue

        for line_number, (l_row_idx, l_row_data) in enumerate(so_lines, start=1):
            l_status = str(l_row_data.get("Status", "")).strip().upper()

            if l_status == "DONE":
                print(f"    ⏭️  Skipping line (already DONE): {so_num} / line {line_number}")
                continue

            explicit_line = iv(l_row_data, "Line Number", 0)
            line_no = explicit_line if explicit_line > 0 else line_number

            print(f"\n  📝 Line {line_no}: item={sv(l_row_data,'Item Code')} qty={fv(l_row_data,'Quantity Ordered')} price={fv(l_row_data,'List Price')} disc={fv(l_row_data,'Discount')} net={fv(l_row_data,'Net Price')}")

            try:
                success, error_msg = create_line(domain, so_num, l_row_data, line_no, token_ref)

                if success:
                    mark_row(ws_l, l_row_idx, l_status_col, l_error_col, "DONE", "")
                    l_success += 1
                else:
                    mark_row(ws_l, l_row_idx, l_status_col, l_error_col, "ERROR", error_msg, RED_FILL)
                    l_fail += 1
                    print(f"    ❌ Line {line_no} failed: {error_msg}")

            except Exception as e:
                mark_row(ws_l, l_row_idx, l_status_col, l_error_col, "ERROR", str(e), RED_FILL)
                l_fail += 1
                print(f"    ⚠️  Line {line_no} exception: {e}")

            wb.save(file_path)
            time.sleep(0.2)

    print(f"\n{'='*60}")
    print(f"📊 SUMMARY")
    print(f"   Headers — ✅ {h_success} success / ❌ {h_fail} failed")
    print(f"   Lines   — ✅ {l_success} success / ❌ {l_fail} failed")
    print(f"{'='*60}")

    return h_success, h_fail, l_success, l_fail


# ---------------------------------------------------------------------------
# Standalone testing mode
# ---------------------------------------------------------------------------
def run_standalone():

    import sys
    import os

    #Temporary fix for standalone execution in cmd
    ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    
    sys.path.append(ROOT_DIR)

    from config import CONFIG

    folder = os.path.abspath(os.path.join(ROOT_DIR, CONFIG["folders"]["sales_order"]))

    if not os.path.exists(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)
    
    xlsx_files = [f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")]

    if not xlsx_files:
        print(f"ERROR: No .xlsx file found in folder: {folder}")
        sys.exit(1)

    total_processed = 0
    total_failed   = 0

    for file_name in xlsx_files:

        file_path = os.path.join(folder, file_name)

        print(f"Validating: {file_path} ...")

        success_count, fail_count = run(file_path)

        print("\nValidation Summary")
        print("------------------------------")
        print(f"Rows processed : {res1ult['rows_processed']}")
        print(f"Rows passed    : {result['rows_passed']}")
        print(f"Rows failed    : {result['rows_failed']}")
        print(f"Rows skipped   : {result['rows_skipped']}")

        total_processed += result["rows_processed"]
        total_failed    += result["rows_failed"]

        if result["has_errors"]:
            print("\nValidation FAILED")
        else:
            print("\nValidation PASSED — all rows are READY for Loading.")

    #outside loop    
    sys.exit(0 if total_failed == 0 else 1)

if __name__ == "__main__":
    run_standalone()