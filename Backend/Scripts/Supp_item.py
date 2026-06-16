import openpyxl
import requests
import time
import sys
import os
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

# ── Path / config ──────────────────────────────────────────────────────────
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT_DIR)
from config import CONFIG

UPLOAD_PATH = "/api/erp/purchasing/supplierItems?viewUri=urn:be:com.qad.purchasing.setup.ISupplierItem"

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)


# =============================================================================
# 1. AUTHENTICATION
# =============================================================================

class _TokenExpired(Exception):
    pass


def _fetch_token() -> str:
    url  = f"{CONFIG['qad']['base_url']}/oauth/token"
    resp = requests.post(url, data=CONFIG["qad"]["auth"], timeout=30)
    resp.raise_for_status()
    token = resp.json().get("access_token")
    if not token:
        raise RuntimeError("OAuth response did not contain access_token")
    return token


class TokenManager:
    """Holds one token for the run; refreshes on demand (401)."""

    def __init__(self):
        self._token: str | None = None

    def get(self) -> str:
        if self._token is None:
            self._token = _fetch_token()
        return self._token

    def refresh(self) -> str:
        self._token = _fetch_token()
        return self._token


# =============================================================================
# 2. DATE HELPER
# =============================================================================

def format_date(val):
    if isinstance(val, (int, float)):
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(val))).strftime("%Y-%m-%dT00:00:00.000Z")
    elif isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT00:00:00.000Z")
    return str(val).strip()


# =============================================================================
# 3. WORKBOOK HELPERS
# =============================================================================

def ensure_columns(ws):
    header_row = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in ws[1]
    ]
    for name in ("Status", "Error"):
        if name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=name)
            header_row.append(name)
    return header_row, header_row.index("Status") + 1, header_row.index("Error") + 1


def mark_error(ws, row_idx, status_col, error_col, error_msg):
    ws.cell(row=row_idx, column=1).fill = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="")
    ws.cell(row=row_idx, column=error_col, value=error_msg)
    ws.cell(row=row_idx, column=error_col).fill = RED_FILL


def mark_success(ws, row_idx, status_col, error_col):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col, value="")


# =============================================================================
# 4. FILE RENAME HELPERS
# =============================================================================

def rename_error(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if not name.startswith("error_"):
        new_path = os.path.join(folder, "error_" + name)
        os.rename(file_path, new_path)
        return new_path
    return file_path


def rename_restore(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if name.startswith("error_"):
        new_path = os.path.join(folder, name[len("error_"):])
        os.rename(file_path, new_path)
        return new_path
    return file_path


# =============================================================================
# 5. API
# =============================================================================

def parse_errors(resp_json):
    errors = resp_json.get("submitResult", {}).get("errors", [])
    msgs = []
    for e in errors:
        msg   = e.get("message", "Unknown error")
        field = e.get("fieldName") or ""
        if "already exists" in msg.lower():
            msgs.append("Duplicate record")
        elif field:
            msgs.append(f"Field '{field}': {msg}")
        else:
            msgs.append(msg)
    return "; ".join(msgs) if msgs else "Failed"


def post_supplier_item(payload: dict, token: str) -> tuple[bool, str]:
    """
    POST one supplier item to QAD.
    Returns (success, error_msg). Raises _TokenExpired on 401.
    """
    url = f"{CONFIG['qad']['base_url']}{UPLOAD_PATH}"

    resp = requests.post(
        url,
        json=payload,
        headers={
            "Content-Type":  "application/json",
            "Authorization": f"Bearer {token}",
        },
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    resp_json = resp.json()

    if resp.status_code == 200 and resp_json.get("submitResult", {}).get("success"):
        return True, ""

    return False, parse_errors(resp_json)


# =============================================================================
# 6. SINGLE-FILE PROCESSOR
# =============================================================================

def process_file(file_path: str, tm: TokenManager) -> tuple[int, int]:
    print(f"\n{'─'*55}")
    print(f"  📄 {os.path.basename(file_path)}")
    print(f"{'─'*55}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_row, status_col, error_col = ensure_columns(ws)

    success_count = skip_count = fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        if str(row_data.get("Status", "")).strip().upper() == "DONE":
            si = str(row_data.get("Supplier Item", "")).strip()
            print(f"  ⏭  {si or f'row {row_idx}'} — already DONE")
            skip_count += 1
            continue

        domain    = str(row_data.get("Domain Code",     "")).strip()
        item      = str(row_data.get("Item Code",       "")).strip()
        supplier  = str(row_data.get("Supplier Code",   "")).strip()
        supp_item = str(row_data.get("Supplier Item",   "")).strip()
        uom       = str(row_data.get("Unit of Measure", "")).strip()
        price     = row_data.get("Quote Price")
        date      = row_data.get("Quote Date")

        missing = []
        if not domain:    missing.append("Domain Code")
        if not item:      missing.append("Item Code")
        if not supplier:  missing.append("Supplier Code")
        if not supp_item: missing.append("Supplier Item")
        if not uom:       missing.append("Unit of Measure")
        if not price:     missing.append("Quote Price")
        if not date:      missing.append("Quote Date")

        if missing:
            msg = f"Missing required field(s): {', '.join(missing)}"
            mark_error(ws, row_idx, status_col, error_col, msg)
            print(f"  ✘  row {row_idx} — {msg}")
            fail_count += 1
            wb.save(file_path)
            continue

        uri = (
            f"urn:be:com.qad.purchasing.setup.ISupplierItem"
            f":{domain}.{item}.{supplier}.{supp_item}"
        )

        item_payload = {
            "uri":           uri,
            "domainCode":    domain,
            "itemCode":      item,
            "supplierCode":  supplier,
            "supplierItem":  supp_item,
            "quoteDate":     format_date(date),
            "quotePrice":    price,
            "unitOfMeasure": uom,
            "dataOperation": "A",
            "currencyCode":  "INR",
        }

        pl = str(row_data.get("Price List", "")).strip()
        if pl:
            item_payload["priceList"] = pl

        payload = {"supplierItems": [item_payload]}

        # ── POST with one token-refresh retry ─────────────────────────────
        success   = False
        error_msg = ""

        for attempt in range(2):
            try:
                success, error_msg = post_supplier_item(payload, tm.get())
                break
            except _TokenExpired:
                if attempt == 0:
                    tm.refresh()
                    continue
                error_msg = "Token refresh failed — unauthorised"
                break
            except requests.RequestException as e:
                error_msg = f"Network error: {e}"
                break

        if success:
            mark_success(ws, row_idx, status_col, error_col)
            print(f"  ✔  Supplier Item: {supp_item} / Item: {item} / Supplier: {supplier}")
            success_count += 1
        else:
            mark_error(ws, row_idx, status_col, error_col, error_msg)
            print(f"  ✘  Supplier Item: {supp_item} — {error_msg}")
            fail_count += 1

        wb.save(file_path)
        time.sleep(0.1)

    if fail_count > 0:
        file_path = rename_error(file_path)
        print(f"\n  ⚠  Errors found — file renamed to: {os.path.basename(file_path)}")
    elif success_count > 0 and fail_count == 0:
        file_path = rename_restore(file_path)

    print(f"\n  📊 Success: {success_count} | Skipped: {skip_count} | Failed: {fail_count}")
    return success_count, fail_count


# =============================================================================
# 7. ORCHESTRATOR
# =============================================================================

def run(folder_path: str) -> tuple[int, int]:
    folder = os.path.abspath(folder_path)

    if not os.path.exists(folder):
        print(f"❌ Folder not found: {folder}")
        raise RuntimeError(f"Folder not found: {folder}")

    files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not files:
        print("⚠  No .xlsx files found in Supplier_Item/")
        raise RuntimeError(f"No .xlsx files found in: {folder}")

    tm = TokenManager()

    total_s = total_f = 0
    for f in files:
        s, fail = process_file(os.path.join(folder, f), tm)
        total_s += s
        total_f += fail

    print(f"\n{'═'*55}")
    print(f"  TOTAL — Success: {total_s} | Failed: {total_f}")
    print(f"{'═'*55}")
    return total_s, total_f


# =============================================================================
# 8. ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    folder = os.path.abspath(
        os.path.join(ROOT_DIR, CONFIG["folders"]["supplier_item"])
    )
    ok, fail = run(folder)
    sys.exit(0 if fail == 0 else 1)