from datetime import datetime
import openpyxl
import requests
import time
import sys
import os
from openpyxl.styles import PatternFill

# ==========================================
# AUTH SETUP
# ==========================================
TOKEN_URL  = "https://cat5-devl.adaptive.qad.com/clouderp/oauth/token"
UPLOAD_URL = "https://cat5-devl.adaptive.qad.com/clouderp/api/erp/productionOrderMasters"

AUTH_PARAMS = {
    "client_id": "ff727cb2cd15c29eed146dcd587a62ed",
    "username": "demo",
    "password": "qad",
    "grant_type": "password"
}

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)


# ==========================================
# HELPERS
# ==========================================

def get_new_token():
    try:
        response = requests.post(TOKEN_URL, params=AUTH_PARAMS)
        response.raise_for_status()
        token = response.json().get("access_token")
        if token:
            return token
        print("❌ Failed to obtain access token.")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Token error: {e}")
        sys.exit(1)


def to_bool(val):
    return str(val).strip().lower() in ("true", "1", "yes")


def to_iso_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%dT%H:%M:%S.000Z')
    elif val and isinstance(val, str):
        try:
            dt = datetime.strptime(val, '%Y-%m-%d %H:%M:%S')
            return dt.strftime('%Y-%m-%dT%H:%M:%S.000Z')
        except ValueError:
            return val
    return val


def ensure_columns(ws):
    header_row = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
    for name in ("Status", "Error"):
        if name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=name)
            header_row.append(name)
    return header_row, header_row.index("Status") + 1, header_row.index("Error") + 1


def mark_error(ws, row_idx, status_col, error_col, error_msg):
    ws.cell(row=row_idx, column=1).fill = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="")
    ws.cell(row=row_idx, column=error_col, value=error_msg)
    ws.cell(row=row_idx, column=error_col).fill = RED_FILL


def mark_success(ws, row_idx, status_col, error_col):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col, value="")


def rename_error(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if not name.startswith("error_"):
        new_path = os.path.join(folder, "error_" + name)
        os.rename(file_path, new_path)
        return new_path
    return file_path


def rename_restore(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if name.startswith("error_"):
        new_path = os.path.join(folder, name[len("error_"):])
        os.rename(file_path, new_path)
        return new_path
    return file_path


# ==========================================
# MAIN
# ==========================================

def run(file_path):
    print(f"\n{'─'*55}")
    print(f"  📄 {os.path.basename(file_path)}")
    print(f"{'─'*55}")

    token = get_new_token()
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_row, status_col, error_col = ensure_columns(ws)

    success_count = skip_count = fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        if str(row_data.get("Status", "")).strip().upper() == "DONE":
            wo = str(row_data.get("workOrderNumber", "")).strip()
            print(f"  ⏭  {wo or f'row {row_idx}'} — already DONE")
            skip_count += 1
            continue

        domain_code      = str(row_data.get("domainCode", "")).strip()
        work_order_num   = str(row_data.get("workOrderNumber", "")).strip()
        work_order_id    = str(row_data.get("workOrderID", "")).strip()
        item_code        = str(row_data.get("itemCode", "")).strip()
        site_code        = str(row_data.get("siteCode", "")).strip()
        work_order_type  = str(row_data.get("workOrderType", "")).strip()

        missing = []
        if not work_order_num: missing.append("workOrderNumber")
        if not work_order_id:  missing.append("workOrderID")

        if missing:
            msg = f"Missing required field(s): {', '.join(missing)}"
            mark_error(ws, row_idx, status_col, error_col, msg)
            print(f"  ✘  row {row_idx} — {msg}")
            fail_count += 1
            continue

        uri = f"urn:be:com.qad.pushproduction.workorder.IWorkOrderMaster:{domain_code}.{work_order_num}.{work_order_id}"

        payload = {
            "supplementaryMessages": [],
            "workOrderMasters": [{
                "uri":                              uri,
                "domainCode":                       domain_code,
                "siteCode":                         site_code,
                "workOrderNumber":                  work_order_num,
                "workOrderID":                      work_order_id,
                "workOrderType":                    work_order_type,
                "itemCode":                         item_code,
                "itemDescription1":                 row_data.get("itemDescription1"),
                "statusCode":                       row_data.get("statusCode"),
                "manufacturingTypeCode":            row_data.get("manufacturingTypeCode"),
                "quantityOrdered":                  row_data.get("quantityOrdered"),
                "quantityExpectedToBeCompleted":    row_data.get("quantityExpectedToBeCompleted"),
                "orderDate":                        to_iso_date(row_data.get("orderDate")),
                "dueDate":                          to_iso_date(row_data.get("dueDate")),
                "needDate":                         to_iso_date(row_data.get("needDate")),
                "releaseDate":                      to_iso_date(row_data.get("releaseDate")),
                "routingCode":                      row_data.get("routingCode"),
                "productionLineCode":               row_data.get("productionLineCode"),
                "productionRatePerHour":            row_data.get("productionRatePerHour"),
                "numberOfProductionLines":          row_data.get("numberOfProductionLines"),
                "unitOfMeasure":                    row_data.get("unitOfMeasure"),
                "runCrewSize":                      row_data.get("runCrewSize"),
                "queuePercent":                     row_data.get("queuePercent"),
                "yieldPercent":                     row_data.get("yieldPercent"),
                "isExplodeBill":                    to_bool(row_data.get("isExplodeBill")),
                "isAutoReschedule":                 to_bool(row_data.get("isAutoReschedule")),
                "floorStockAcct":                   row_data.get("floorStockAcct"),
                "floorStockSubAcct":                row_data.get("floorStockSubAcct"),
                "workInProcessAcct":                row_data.get("workInProcessAcct"),
                "workInProcessSubAcct":             row_data.get("workInProcessSubAcct"),
                "dataOperation":                    "U"
            }]
        }

        try:
            response = requests.post(
                UPLOAD_URL, json=payload,
                headers={"Content-Type": "application/json", "Authorization": f"Bearer {token}"}
            )

            if response.status_code == 401:
                token = get_new_token()
                response = requests.post(
                    UPLOAD_URL, json=payload,
                    headers={"Content-Type": "application/json", "Authorization": f"Bearer {token}"}
                )

            resp_json = response.json()
            submit = resp_json.get("submitResult", {})

            if submit.get("success") is True:
                mark_success(ws, row_idx, status_col, error_col)
                print(f"  ✔  Work Order: {work_order_num}")
                success_count += 1
            else:
                errors = submit.get("errors", [])
                msgs = []
                for e in errors:
                    msg = e.get("message", "")
                    if "already exists" in msg.lower():
                        msgs.append(f"Duplicate — Work Order '{work_order_num}' already exists")
                    elif e.get("fieldName"):
                        msgs.append(f"Field '{e['fieldName']}': {msg}")
                    else:
                        msgs.append(msg)
                error_msg = "; ".join(msgs) or f"HTTP {response.status_code}"
                mark_error(ws, row_idx, status_col, error_col, error_msg)
                print(f"  ✘  Work Order: {work_order_num} — {error_msg}")
                fail_count += 1

        except Exception as e:
            mark_error(ws, row_idx, status_col, error_col, str(e))
            print(f"  ✘  Work Order: {work_order_num} — Connection error: {e}")
            fail_count += 1

        time.sleep(0.2)

    wb.save(file_path)

    if fail_count > 0:
        file_path = rename_error(file_path)
        print(f"\n  ⚠  Errors found — file renamed to: {os.path.basename(file_path)}")
    elif success_count > 0 and fail_count == 0:
        file_path = rename_restore(file_path)

    print(f"\n  📊 Success: {success_count} | Skipped: {skip_count} | Failed: {fail_count}")
    return success_count, fail_count


# ==========================================
# ENTRY POINT
# ==========================================
if __name__ == "__main__":
    base   = os.path.dirname(os.path.abspath(__file__))
    folder = os.path.join(base, "..", "ProductionOrder")

    if not os.path.exists(folder):
        print(f"❌ Folder not found: {folder}")
        sys.exit(1)

    files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not files:
        print("⚠  No .xlsx files found in ProductionOrder/")
        sys.exit(0)

    total_s = total_f = 0
    for f in files:
        s, fail = run(os.path.join(folder, f))
        total_s += s
        total_f += fail

    print(f"\n{'═'*55}")
    print(f"  TOTAL — Success: {total_s} | Failed: {total_f}")
    print(f"{'═'*55}")
    sys.exit(0 if total_f == 0 else 1)