"""
BR_load.py
----------
Loads Business Relation rows from all .xlsx files in the configured BR folder
into QAD via the businessRelationV2s API.

Behaviour
---------
- Fetches one OAuth token per run; refreshes only on 401
- Skips rows with Status = DONE
- Lightweight mandatory-field check before any API call
- Success check: submitResult.success == True
- Error messages extracted from submitResult.errors[].message
- On success  → clear all red fills, clear Error, set Status = DONE
- On failure  → red-fill first col + bad cols + Error col, log message, set Status = ERROR
- Processes every row regardless of individual failures
- CREATE only — no update operation
- Returns (ok_count, fail_count)
"""

import os
import sys
import time
import requests
import openpyxl
from openpyxl.styles import PatternFill

# ── Path / config ─────────────────────────────────────────────────────────
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT_DIR)
from config import CONFIG

# ── Fill constants ────────────────────────────────────────────────────────
RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

# ── Mandatory columns ─────────────────────────────────────────────────────
MANDATORY_COLUMNS = [
    "Business Relation",
    "Name",
    "Active",
    "Language",
    "Country",
]


# =============================================================================
# 1. AUTHENTICATION
# =============================================================================

def _fetch_token() -> str:
    url  = f"{CONFIG['qad']['base_url']}/oauth/token"
    resp = requests.post(url, data=CONFIG["qad"]["auth"], timeout=30)
    resp.raise_for_status()
    token = resp.json().get("access_token")
    if not token:
        raise RuntimeError("OAuth response did not contain access_token")
    return token


class TokenManager:
    """Holds one token for the run; refreshes on demand (401)."""

    def __init__(self):
        self._token: str | None = None

    def get(self) -> str:
        if self._token is None:
            self._token = _fetch_token()
        return self._token

    def refresh(self) -> str:
        self._token = _fetch_token()
        return self._token


# =============================================================================
# 2. PAYLOAD BUILDER
# =============================================================================

def build_payload(row: dict) -> dict:
    def val(col: str) -> str:
        v = row.get(col)
        return str(v).strip() if v is not None else ""

    br_code = val("Business Relation")
    name    = val("Name")
    uri     = f"urn:service:com.qad.base.address.IEmployee-BusinessRelationV2s:{br_code}"

    return {
        "supplementaryMessages": [],
        "businessRelationV2s": [
            {
                "uri":                          uri,
                "businessRelationCode":         br_code,
                "businessRelationID":           0,
                "businessRelationName1":        name,
                "businessRelationName2":        "",
                "businessRelationName3":        "",
                "businessRelationSearchName":   val("Search Name"),
                "isActive":                     val("Active").lower() == "yes",
                "languageCode":                 val("Language"),
                "headOfficeAddressTypeCode":    val("Address Type") or "HEADOFFICE",
                "headOfficeAddressName":        name,
                "headOfficeAddressSearchName":  val("Search Name"),
                "headOfficeStreet1":            val("Address 1"),
                "headOfficeStreet2":            val("Address 2"),
                "headOfficeStreet3":            "",
                "headOfficeCity":               val("City"),
                "headOfficeStateCode":          val("State"),
                "headOfficeZipCode":            val("Postal Code"),
                "headOfficeCountryCode":        val("Country"),
                "headOfficeTaxZone":            val("Tax Zone"),
                "changeStatus":                 "2",
                "isPreDefaulted":               True,
                "addressV2s":                   [],
                "busRelationSafDefaultV2s":     [],
            }
        ],
    }


# =============================================================================
# 3. API
# =============================================================================

class _TokenExpired(Exception):
    pass


def post_br(payload: dict, token: str) -> tuple[bool, str]:
    """
    POST BR payload to QAD.
    Returns (success, error_msg). Raises _TokenExpired on 401.
    """
    url = (
        f"{CONFIG['qad']['base_url']}/api/erp/businessRelationV2s"
        f"?viewUri=urn:be:com.qad.base.address.IBusinessRelationV2"
    )

    resp = requests.post(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/json",
        },
        json=payload,
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    resp_json = resp.json()
    submit    = resp_json.get("submitResult", {})

    if submit.get("success") is True:
        return True, ""

    errors    = submit.get("errors", [])

    error_messages = []
    bad_cols = []

    for e in errors:
        msg =  e.get("message", "").strip()

        #extracting field name
        field_name = (
            e.get("fieldName")
            or e.get("field")
            or e.get("name")
            or ""
        ).strip()

        if field_name:
            print(f"{field_name} -> {msg}")
            bad_cols.append(field_name)
        else:
            print(msg)
        
        if msg:
            error_messages.append(msg)
    
    
    error_msg = "; ".join(error_messages) or (
        f"HTTP {resp.status_code} - submitResult.success was not True"
    )

    return False, (error_msg, bad_cols)


# =============================================================================
# 4. WORKBOOK HELPERS
# =============================================================================

def _mark_done(ws, row_idx: int, status_col: int, error_col: int):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col,  value="")


def _mark_error(
    ws,
    row_idx:    int,
    status_col: int,
    error_col:  int,
    header_row: list,
    bad_cols:   list,
    error_msg:  str,
):
    
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL

    

    for col_name in bad_cols:
        
        #API -> excel names
        mapping = {
            "BusinessRelationCode": "Business Relation",
            "HeadOfficeCity": "City",
            "HeadOfficeStateCode": "State",
            "HeadOfficeZipCode": "Postal Code",
            "HeadOfficeCountryCode": "Country",
        }

        excel_col = mapping.get(col_name, col_name)

        if excel_col in header_row:
            col_idx = header_row.index(excel_col) + 1
            ws.cell(row = row_idx, column=col_idx).fill = RED_FILL

    ws.cell(row=row_idx, column=1).fill         = RED_FILL
    ws.cell(row=row_idx, column=status_col).fill = RED_FILL
    ws.cell(row=row_idx, column=error_col).fill  = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="ERROR")
    ws.cell(row=row_idx, column=error_col,  value=error_msg)


# =============================================================================
# 5. MANDATORY FIELD CHECK
# =============================================================================

def _check_mandatory(row_data: dict) -> list[str]:
    missing = []
    for col in MANDATORY_COLUMNS:
        v = row_data.get(col)
        if v is None or str(v).strip() == "" or str(v).strip().lower() == "none":
            missing.append(col)
    return missing


# =============================================================================
# 6. SINGLE-FILE PROCESSOR
# =============================================================================

def process_file(file_path: str, tm: TokenManager) -> tuple[int, int]:
    """Open one workbook, process every row, return (ok_count, fail_count)."""

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_row = [
        str(c.value).strip() if c.value is not None else ""
        for c in ws[1]
    ]

    for col_name in ("Status", "Error"):
        if col_name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=col_name)
            header_row.append(col_name)

    status_col = header_row.index("Status") + 1
    error_col  = header_row.index("Error")  + 1

    ok_count   = 0
    fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):

        row_values = [cell.value for cell in row]

        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        status = str(row_data.get("Status", "")).strip().upper()

        # Skip already completed rows
        if status == "DONE":
            continue

        # Mandatory check
        missing = _check_mandatory(row_data)
        if missing:
            fail_count += 1
            _mark_error(
                ws, row_idx, status_col, error_col, header_row,
                missing,
                f"Missing mandatory fields: {', '.join(missing)}",
            )
            wb.save(file_path)
            continue

        # ── Build payload and POST ─────────────────────────────────────────
        payload   = build_payload(row_data)
        success   = False
        error_msg = ""
        bad_cols  = []

        for attempt in range(2):
            try:
                result = post_br(payload, tm.get())

                if result[0]:
                    success = True
                    error_msg = ""
                    bad_cols = []
                else:
                    success = False
                    error_msg, bad_cols = result[1]
                break
            except _TokenExpired:
                if attempt == 0:
                    tm.refresh()
                    continue
                error_msg = "Token refresh failed — unauthorised"
                break
            except requests.RequestException as e:
                error_msg = f"Network error: {e}"
                break

        if success:
            ok_count += 1
            _mark_done(ws, row_idx, status_col, error_col)
        else:
            fail_count += 1
            _mark_error(ws, row_idx, status_col, error_col, header_row, bad_cols, error_msg)

        wb.save(file_path)
        time.sleep(0.1)

    return ok_count, fail_count


# =============================================================================
# 7. ORCHESTRATOR
# =============================================================================

def run(folder_path: str) -> tuple[int, int]:
    """
    Scan folder_path for .xlsx files, process each one, return (total_ok, total_fail).
    This is the single entry point used by both main.py and __main__.
    """
    folder = os.path.abspath(folder_path)

    if not os.path.exists(folder):
        print(f"ERROR: Folder not found: {folder}")
        raise RuntimeError(f"Folder not found: {folder}")

    xlsx_files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not xlsx_files:
        print(f"ERROR: No .xlsx files found in: {folder}")
        raise RuntimeError(f"No .xlsx files found in: {folder}")

    tm = TokenManager()

    total_ok   = 0
    total_fail = 0

    for file_name in xlsx_files:
        file_path = os.path.join(folder, file_name)
        print(f"Loading : {file_path}")

        ok, fail   = process_file(file_path, tm)
        total_ok   += ok
        total_fail += fail

        print("\nLoad Summary")
        print("-" * 40)
        print(f"  Rows loaded successfully : {ok}")
        print(f"  Rows failed              : {fail}")
        if fail == 0:
            print("  Result : ALL ROWS LOADED SUCCESSFULLY ✓")
        else:
            print("  Result : COMPLETED WITH ERRORS — fix red rows and re-run")
        print()

    print("=" * 40)
    print(f"  Total loaded : {total_ok}")
    print(f"  Total failed : {total_fail}")
    print("=" * 40)

    return total_ok, total_fail


# =============================================================================
# 8. ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    folder = os.path.abspath(
        os.path.join(ROOT_DIR, CONFIG["folders"]["BR"])
    )
    ok, fail = run(folder)
    sys.exit(0 if fail == 0 else 1)