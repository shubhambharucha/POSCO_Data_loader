import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

# All columns present in workorder_mandatory.xlsx, mapped to their types
FIELD_RULES = {
    "domainCode":                       {"type": "character", "max_len": 8},
    "siteCode":                         {"type": "character", "max_len": 8},
    "workOrderNumber":                  {"type": "character", "max_len": 8},
    "workOrderID":                      {"type": "character", "max_len": 8},
    "workOrderType":                    {"type": "character", "max_len": 1},
    "itemCode":                         {"type": "character", "max_len": 18},
    "itemDescription1":                 {"type": "character", "max_len": 24},
    "statusCode":                       {"type": "character", "max_len": 2},
    "manufacturingTypeCode":            {"type": "character", "max_len": 2},
    "quantityOrdered":                  {"type": "decimal",   "min": 0},
    "quantityExpectedToBeCompleted":    {"type": "decimal",   "min": 0},
    "orderDate":                        {"type": "date"},
    "dueDate":                          {"type": "date"},
    "needDate":                         {"type": "date"},
    "releaseDate":                      {"type": "date"},
    "routingCode":                      {"type": "character", "max_len": 8},
    "productionLineCode":               {"type": "character", "max_len": 8},
    "productionRatePerHour":            {"type": "decimal",   "min": 0},
    "numberOfProductionLines":          {"type": "integer"},
    "unitOfMeasure":                    {"type": "character", "max_len": 2},
    "runCrewSize":                      {"type": "decimal",   "min": 0},
    "queuePercent":                     {"type": "decimal",   "min": 0},
    "yieldPercent":                     {"type": "decimal",   "min": 0},
    "isExplodeBill":                    {"type": "logical"},
    "isAutoReschedule":                 {"type": "logical"},
    "floorStockAcct":                   {"type": "character", "max_len": 8},
    "floorStockSubAcct":                {"type": "character", "max_len": 8},
    "workInProcessAcct":                {"type": "character", "max_len": 8},
    "workInProcessSubAcct":             {"type": "character", "max_len": 8},
    "dataOperation":                    {"type": "character", "max_len": 1},
}

ENTITY_COL    = "domainCode"
VALID_LOGICAL = {"true", "false", "yes", "no", "1", "0"}


def _check_char(value, max_len):
    if value is None or str(value).strip() == "" or str(value).strip().lower() == "none":
        return "empty"
    if len(str(value).strip()) > max_len:
        return f"max {max_len} chars (got {len(str(value).strip())})"
    return None


def _check_decimal(value, min_val=0):
    if value is None or str(value).strip() == "":
        return "empty"
    try:
        if float(value) < min_val:
            return f"must be >= {min_val}"
    except (ValueError, TypeError):
        return "invalid number"
    return None


def _check_integer(value):
    if value is None or str(value).strip() == "":
        return "empty"
    try:
        int(float(value))
    except (ValueError, TypeError):
        return "invalid integer"
    return None


def _check_date(value):
    if value is None or str(value).strip() == "":
        return "empty"
    if isinstance(value, datetime):
        return None
    try:
        float(value)
        return None
    except (ValueError, TypeError):
        pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S.000Z"):
        try:
            datetime.strptime(str(value).strip(), fmt)
            return None
        except ValueError:
            continue
    return "invalid date"


def _check_logical(value):
    if value is None or str(value).strip() == "":
        return "empty"
    if str(value).strip().lower() not in VALID_LOGICAL:
        return f"must be True/False (got '{value}')"
    return None


def validate(file_path):
    wb = openpyxl.load_workbook(file_path)

    # Sheet is named "WO Mandatory Fields"
    sheet_name = "WO Mandatory Fields"
    if sheet_name not in wb.sheetnames:
        # Fall back to active sheet if name differs
        ws = wb.active
    else:
        ws = wb[sheet_name]

    header_row = [cell.value for cell in ws[1]]

    if "Status" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Status")
        header_row.append("Status")
    if "Error" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Error")
        header_row.append("Error")

    status_col_idx = header_row.index("Status") + 1
    error_col_idx  = header_row.index("Error")  + 1
    entity_col_idx = header_row.index(ENTITY_COL) + 1

    has_errors  = False
    error_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]

        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        current_status = str(row_data.get("Status", "")).strip().upper()
        if current_status in ("DONE", "READY"):
            continue

        row_errors      = []
        error_cell_idxs = []

        for col_name, rule in FIELD_RULES.items():
            if col_name not in header_row:
                continue

            col_idx = header_row.index(col_name) + 1
            value   = row_data.get(col_name)

            if rule["type"] == "character":
                err = _check_char(value, rule["max_len"])
            elif rule["type"] == "decimal":
                err = _check_decimal(value, rule.get("min", 0))
            elif rule["type"] == "integer":
                err = _check_integer(value)
            elif rule["type"] == "date":
                err = _check_date(value)
            elif rule["type"] == "logical":
                err = _check_logical(value)
            else:
                err = None

            if err:
                row_errors.append(f"{col_name}: {err}")
                error_cell_idxs.append(col_idx)

        for cell in ws[row_idx]:
            cell.fill = CLEAR_FILL

        if row_errors:
            has_errors  = True
            error_count += 1
            error_msg   = "; ".join(row_errors)

            ws.cell(row=row_idx, column=status_col_idx, value="")
            ws.cell(row=row_idx, column=entity_col_idx).fill = RED_FILL
            for cidx in error_cell_idxs:
                ws.cell(row=row_idx, column=cidx).fill = RED_FILL
            ws.cell(row=row_idx, column=error_col_idx, value=error_msg).fill = RED_FILL

        else:
            ws.cell(row=row_idx, column=status_col_idx, value="READY")
            ws.cell(row=row_idx, column=error_col_idx,  value="")

    wb.save(file_path)
    return has_errors, error_count