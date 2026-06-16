import openpyxl
import requests
import time
import sys
import os
from datetime import datetime
from openpyxl.styles import PatternFill

# ==========================================
# AUTH SETUP
# ==========================================
TOKEN_URL  = "https://cat5-devl.adaptive.qad.com/clouderp/oauth/token"
UPLOAD_URL = (
    "https://cat5-devl.adaptive.qad.com/clouderp/api/erp/supplierPriceListV2s"
    "?viewUri=urn:be:com.qad.base.item.ISupplierPriceListV2"
)

AUTH_PARAMS = {
    "client_id": "afb97fd221925b87f01489aeb0e02e81",
    "username":  "demo",
    "password":  "qad",
    "grant_type": "password"
}

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)


# ==========================================
# HELPERS
# ==========================================

def get_new_token():
    try:
        resp = requests.post(TOKEN_URL, params=AUTH_PARAMS, timeout=30)
        resp.raise_for_status()
        token = resp.json().get("access_token")
        if token:
            return token
        print("❌ Failed to obtain access token.")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Token error: {e}")
        sys.exit(1)


def parse_date(value):
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%dT00:00:00.000Z")
    s = str(value).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT00:00:00.000Z")
        except ValueError:
            continue
    return s


def sv(value, default=""):
    return str(value).strip() if value is not None else default

def fv(value, default=0.0):
    try:
        return float(value) if value not in (None, "") else default
    except (ValueError, TypeError):
        return default

def iv(value, default=0):
    try:
        return int(value) if value not in (None, "") else default
    except (ValueError, TypeError):
        return default

def bv(value, default=False):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes") if value else default


def ensure_columns(ws):
    header_row = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
    for name in ("Status", "Error"):
        if name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=name)
            header_row.append(name)
    return header_row, header_row.index("Status") + 1, header_row.index("Error") + 1


def mark_error(ws, row_idx, status_col, error_col, error_msg):
    ws.cell(row=row_idx, column=1).fill = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="")
    ws.cell(row=row_idx, column=error_col, value=error_msg)
    ws.cell(row=row_idx, column=error_col).fill = RED_FILL


def mark_success(ws, row_idx, status_col, error_col):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col, value="")


def rename_error(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if not name.startswith("error_"):
        new_path = os.path.join(folder, "error_" + name)
        os.rename(file_path, new_path)
        return new_path
    return file_path


def rename_restore(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if name.startswith("error_"):
        new_path = os.path.join(folder, name[len("error_"):])
        os.rename(file_path, new_path)
        return new_path
    return file_path


# ==========================================
# MAIN
# ==========================================

def run(file_path):
    print(f"\n{'─'*55}")
    print(f"  📄 {os.path.basename(file_path)}")
    print(f"{'─'*55}")

    current_token = get_new_token()

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"❌ Could not open workbook: {e}")
        return 0, 1

    ws = wb.active
    header_row, status_col, error_col = ensure_columns(ws)

    success_count = skip_count = fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(v for v in row_values if v not in (None, "")):
            continue

        row_data = dict(zip(header_row, row_values))

        if sv(row_data.get("Status")).upper() == "DONE":
            pl = sv(row_data.get("Price List Code"))
            print(f"  ⏭  {pl or f'row {row_idx}'} — already DONE")
            skip_count += 1
            continue

        domain_code     = sv(row_data.get("Domain Code"))
        price_list_code = sv(row_data.get("Price List Code"))
        currency_code   = sv(row_data.get("Currency Code"))
        item_code       = sv(row_data.get("Item Code"))
        start_date_raw  = row_data.get("Start Date")
        expire_date_raw = row_data.get("Expire Date")
        site_code       = sv(row_data.get("Site Code"))

        missing = []
        if not domain_code:     missing.append("Domain Code")
        if not price_list_code: missing.append("Price List Code")
        if not currency_code:   missing.append("Currency Code")
        if not item_code:       missing.append("Item Code")
        if not start_date_raw:  missing.append("Start Date")
        if not expire_date_raw: missing.append("Expire Date")
        if not site_code:       missing.append("Site Code")

        if missing:
            msg = f"Missing required field(s): {', '.join(missing)}"
            mark_error(ws, row_idx, status_col, error_col, msg)
            print(f"  ✘  row {row_idx} — {msg}")
            fail_count += 1
            continue

        start_date             = parse_date(start_date_raw)
        expire_date            = parse_date(expire_date_raw)
        item_list_price        = fv(row_data.get("Item List Price"))
        list_price             = fv(row_data.get("List Price"))
        amount_type            = sv(row_data.get("Amount Type"), default="P") or "P"
        unit_of_measure        = sv(row_data.get("Unit of Measure"))
        price_list_description = sv(row_data.get("Price List Description"))
        product_line           = sv(row_data.get("Product Line"))
        is_tax_included        = bv(row_data.get("Is Tax Included"))
        is_temporary           = bv(row_data.get("Is Temporary"))
        list_classification    = iv(row_data.get("List Classification"), default=1)
        minimum_price          = fv(row_data.get("Minimum Price"))
        maximum_price          = fv(row_data.get("Maximum Price"))
        custom_note            = sv(row_data.get("Custom Note"))

        uri = f"urn:be:com.qad.base.item.ISupplierPriceListV2:{domain_code}..{currency_code}...."

        payload = {
            "supplementaryMessages": [],
            "supplierPriceLists": [{
                "uri":                    uri,
                "amountType":             amount_type,
                "concurrencyHash":        "",
                "currencyCode":           currency_code,
                "currencyDescription":    "",
                "customDecimal0": 0, "customDecimal1": 0, "customDecimal2": 0,
                "customDecimal3": 0, "customDecimal4": 0,
                "customInteger0": 0, "customInteger1": 0, "customInteger2": 0,
                "customInteger3": 0, "customInteger4": 0,
                "customLong0": "", "customLong1": "",
                "customNote":             custom_note,
                "customShort0":  "", "customShort1":  "", "customShort2":  "",
                "customShort3":  "", "customShort4":  "", "customShort5":  "",
                "customShort6":  "", "customShort7":  "", "customShort8":  "",
                "customShort9":  "", "customShort10": "", "customShort11": "",
                "customShort12": "", "customShort13": "", "customShort14": "",
                "customShort15": "", "customShort16": "", "customShort17": "",
                "customShort18": "", "customShort19": "",
                "dataOperation":          "",
                "domainCode":             domain_code,
                "domainCurrency":         currency_code,
                "expireDate":             expire_date,
                "isTaxIncluded":          is_tax_included,
                "isTemporary":            is_temporary,
                "itemCode":               item_code,
                "itemDescription":        "",
                "itemListPrice":          item_list_price,
                "lastModifiedUser":       "",
                "listClassification":     list_classification,
                "listPrice":              list_price,
                "listType":               "SUPPLIER",
                "maximumPrice":           maximum_price,
                "minimumPrice":           minimum_price,
                "pcMstrUser1": "", "pcMstrUser2": "",
                "priceListCode":          price_list_code,
                "priceListDescription":   price_list_description,
                "prodLineDescription":    "",
                "productLine":            product_line,
                "siteCode":               site_code,
                "siteDescription":        "",
                "startDate":              start_date,
                "stockUOM":               unit_of_measure,
                "stockUOMDescription":    "",
                "supplierPriceListTiers": [],
                "thisLevelGLCost":        0,
                "totalGLCost":            0,
                "unitOfMeasure":          unit_of_measure,
                "unitOfMeasureDescription": "",
            }]
        }

        retry = True
        while retry:
            try:
                resp = requests.post(
                    UPLOAD_URL, json=payload,
                    headers={"Content-Type": "application/json", "Authorization": f"Bearer {current_token}"},
                    timeout=30
                )

                if resp.status_code == 401:
                    current_token = get_new_token()
                    continue

                try:
                    resp_json = resp.json()
                except ValueError:
                    resp_json = {}

                submit = resp_json.get("submitResult", {})

                if resp.status_code == 200 and submit.get("success") is True:
                    mark_success(ws, row_idx, status_col, error_col)
                    print(f"  ✔  Price List: {price_list_code} / Item: {item_code}")
                    success_count += 1
                else:
                    errors = submit.get("errors", [])
                    msgs = []
                    for e in errors:
                        msg = e.get("message", "")
                        if "already exists" in msg.lower():
                            msgs.append(f"Duplicate — Price List '{price_list_code}/{item_code}' already exists")
                        elif e.get("fieldName"):
                            msgs.append(f"Field '{e['fieldName']}': {msg}")
                        else:
                            msgs.append(msg)
                    error_msg = "; ".join(msgs) or resp_json.get("message", f"HTTP {resp.status_code}")
                    mark_error(ws, row_idx, status_col, error_col, error_msg)
                    print(f"  ✘  Price List: {price_list_code} / Item: {item_code} — {error_msg}")
                    fail_count += 1

                retry = False

            except requests.exceptions.Timeout:
                msg = "Request timed out after 30s"
                mark_error(ws, row_idx, status_col, error_col, msg)
                print(f"  ✘  Price List: {price_list_code} — {msg}")
                fail_count += 1
                retry = False

            except Exception as exc:
                mark_error(ws, row_idx, status_col, error_col, str(exc))
                print(f"  ✘  Price List: {price_list_code} — Connection error: {exc}")
                fail_count += 1
                retry = False

        try:
            wb.save(file_path)
        except PermissionError:
            print(f"\n❌ Cannot save — file is open in Excel. Close it and press Enter…")
            input()
            wb.save(file_path)

        time.sleep(0.1)

    if fail_count > 0:
        file_path = rename_error(file_path)
        print(f"\n  ⚠  Errors found — file renamed to: {os.path.basename(file_path)}")
    elif success_count > 0 and fail_count == 0:
        file_path = rename_restore(file_path)

    print(f"\n  📊 Success: {success_count} | Skipped: {skip_count} | Failed: {fail_count}")
    return success_count, fail_count


# ==========================================
# ENTRY POINT
# ==========================================
if __name__ == "__main__":
    base   = os.path.dirname(os.path.abspath(__file__))
    folder = os.path.join(base, "..", "SupplierPriceList")

    if not os.path.exists(folder):
        print(f"❌ Folder not found: {folder}")
        sys.exit(1)

    files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not files:
        print("⚠  No .xlsx files found in SupplierPriceList/")
        sys.exit(0)

    total_s = total_f = 0
    for f in files:
        s, fail = run(os.path.join(folder, f))
        total_s += s
        total_f += fail

    print(f"\n{'═'*55}")
    print(f"  TOTAL — Success: {total_s} | Failed: {total_f}")
    print(f"{'═'*55}")
    sys.exit(0 if total_f == 0 else 1)