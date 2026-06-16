import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

# Fields present in Supplier_price_Lists.xlsx — all mandatory
FIELD_RULES = {
    "Domain Code":      {"type": "character", "max_len": 8},
    "Price List Code":  {"type": "character", "max_len": 8},
    "Currency Code":    {"type": "character", "max_len": 3},
    "Item Code":        {"type": "character", "max_len": 18},
    "Start Date":       {"type": "date"},
    "Expire Date":      {"type": "date"},
    "Site Code":        {"type": "character", "max_len": 8},
}

ENTITY_COL = "Domain Code"


def _check_char(value, max_len):
    if value is None or str(value).strip() == "" or str(value).strip().lower() == "none":
        return "empty"
    if len(str(value).strip()) > max_len:
        return f"max {max_len} chars (got {len(str(value).strip())})"
    return None


def _check_date(value):
    if value is None or str(value).strip() == "":
        return "empty"
    if isinstance(value, datetime):
        return None
    try:
        float(value)   # Excel serial date
        return None
    except (ValueError, TypeError):
        pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S.000Z"):
        try:
            datetime.strptime(str(value).strip(), fmt)
            return None
        except ValueError:
            continue
    return "invalid date"


def validate(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]

    if "Status" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Status")
        header_row.append("Status")
    if "Error" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Error")
        header_row.append("Error")

    status_col_idx = header_row.index("Status") + 1
    error_col_idx  = header_row.index("Error")  + 1
    entity_col_idx = header_row.index(ENTITY_COL) + 1

    has_errors  = False
    error_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]

        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        current_status = str(row_data.get("Status", "")).strip().upper()
        if current_status in ("DONE", "READY"):
            continue

        row_errors      = []
        error_cell_idxs = []

        for col_name, rule in FIELD_RULES.items():
            if col_name not in header_row:
                continue

            col_idx = header_row.index(col_name) + 1
            value   = row_data.get(col_name)

            if rule["type"] == "character":
                err = _check_char(value, rule["max_len"])
            elif rule["type"] == "date":
                err = _check_date(value)
            else:
                err = None

            if err:
                row_errors.append(f"{col_name}: {err}")
                error_cell_idxs.append(col_idx)

        for cell in ws[row_idx]:
            cell.fill = CLEAR_FILL

        if row_errors:
            has_errors  = True
            error_count += 1
            error_msg   = "; ".join(row_errors)

            ws.cell(row=row_idx, column=status_col_idx, value="")
            ws.cell(row=row_idx, column=entity_col_idx).fill = RED_FILL
            for cidx in error_cell_idxs:
                ws.cell(row=row_idx, column=cidx).fill = RED_FILL
            ws.cell(row=row_idx, column=error_col_idx, value=error_msg).fill = RED_FILL

        else:
            ws.cell(row=row_idx, column=status_col_idx, value="READY")
            ws.cell(row=row_idx, column=error_col_idx,  value="")

    wb.save(file_path)
    return has_errors, error_count