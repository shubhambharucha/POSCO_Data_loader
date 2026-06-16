"""
Supplier_load.py
----------------
Loads supplier rows from all .xlsx files in the configured supplier folder
into QAD via the supplierV2s API, then updates domain settings via mfgSuppliers.

Behaviour
---------
- Fetches one OAuth token per run; refreshes only on 401
- Skips rows with Status = DONE (for Create operations only)
- Lightweight mandatory-field check before any API call
- Success check: submitResult.success == True
- Error messages extracted from submitResult.errors[].message
- On success  → clear all red fills, clear Error, set Status = DONE
- On failure  → red-fill first col + bad cols + Error col, log message, set Status = ERROR
- Processes every row regardless of individual failures
- Returns (ok_count, fail_count)

CREATE flow (Data Operation = C):
  1. POST to supplierV2s   → creates supplier
  2. GET  mfgSuppliers     → fetch auto-created domain record
  3. PATCH + POST          → update siteCode + daybookSetCode
  If step 3 fails → row marked ERROR with "Supplier created but domain settings failed: ..."

UPDATE flow (Data Operation = U):
  1. GET  supplierV2s      → fetch existing record
  2. Patch editable fields
  3. POST supplierV2s      → update supplier
  (Domain settings NOT touched on UPDATE — one-time setup only)
"""

import os
import sys
import time
import json
import requests
import openpyxl
from openpyxl.styles import PatternFill

# ── Path / config ─────────────────────────────────────────────────────────
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT_DIR)
from config import CONFIG

# ── Fill constants ────────────────────────────────────────────────────────
RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

# ── Mandatory columns (supplier) ──────────────────────────────────────────
MANDATORY_COLUMNS = [
    "Supplier",
    "Shared Set",
    "Business Relation",
    "Active",
    "Currency",
    "Credit Terms",
    "Invoice Status",
    "Invoice Control GL Profile",
    "Credit Note Control GL Profile",
    "Prepayment Control GL Profile",
    "Purchase Account GL Profile",
]

# ── Mandatory columns for domain settings (CREATE only) ───────────────────
MANDATORY_DOMAIN_COLUMNS = [
    "Domain",
    "Site Code",
    "Daybook Set",
]


# =============================================================================
# 1. AUTHENTICATION
# =============================================================================

def _fetch_token() -> str:
    url  = f"{CONFIG['qad']['base_url']}/oauth/token"
    resp = requests.post(url, data=CONFIG["qad"]["auth"], timeout=30)
    resp.raise_for_status()
    token = resp.json().get("access_token")
    if not token:
        raise RuntimeError("OAuth response did not contain access_token")
    return token


class TokenManager:
    """Holds one token for the run; refreshes on demand (401)."""

    def __init__(self):
        self._token: str | None = None

    def get(self) -> str:
        if self._token is None:
            self._token = _fetch_token()
        return self._token

    def refresh(self) -> str:
        self._token = _fetch_token()
        return self._token


# =============================================================================
# 2. PAYLOAD BUILDER  (explicit — every field visible)
# =============================================================================

def build_payload(row: dict) -> dict:
    def val(col: str) -> str:
        v = row.get(col)
        return str(v).strip() if v is not None else ""

    def bool_val(col: str) -> bool:
        return val(col).lower() == "yes"

    def float_val(col: str) -> float:
        return float(val(col) or 0)

    def int_val(col: str) -> int:
        return int(float(val(col) or 0))

    supplier_code = val("Supplier")
    shared_set    = val("Shared Set")
    uri           = f"urn:be:com.qad.base.supplier.ISupplierV2:{shared_set}.{supplier_code}"

    return {
        "supplementaryMessages": [],
        "supplierV2s": [
            {
                # ── URIs / Identity ───────────────────────────────────────────
                "uri":                                      uri,
                "instanceURI":                              uri,
                "supplierCode":                             supplier_code,
                "supplierID":                               0,
                "sharedSetCode":                            shared_set,
                "sharedSetID":                              0,

                # ── Business Relation ─────────────────────────────────────────
                "businessRelationCode":                     val("Business Relation"),
                "businessRelationID":                       0,
                "businessRelationName":                     "",
                "businessRelationName2":                    "",
                "businessRelationName3":                    "",
                "businessRelationConcurrencyHash":          "",
                "isBusinessRelationActive":                 True,
                "isBusinessRelationFieldsEnabled":          True,
                "isBusinessRelationIntercompany":           False,
                "isCreateBusinessRelationRequired":         True,
                "isInternalEntity":                         False,
                "intercompanyCode":                         "",

                # ── Address ───────────────────────────────────────────────────
                "addressID":                                0,
                "addressName":                              val("Name"),
                "addressSearchName":                        val("Search Name"),
                "addressTypeCode":                          val("Address Type"),
                "addressConcurrencyHash":                   "",
                "street1":                                  val("Address 1"),
                "street2":                                  val("Address 2"),
                "street3":                                  "",
                "city":                                     val("City"),
                "cityCode":                                 "",
                "zipCode":                                  val("Postal Code"),
                "stateCode":                                val("State"),
                "stateDescription":                         "",
                "stateTax":                                 val("State Tax"),
                "countryCode":                              val("Country"),
                "countryDescription":                       "",
                "countyCode":                               "",
                "countyDescription":                        "",
                "postalFormat":                             "0",
                "latitude":                                 0,
                "longitude":                                0,
                "isTemporaryAddress":                       False,

                # ── Status / Concurrency ──────────────────────────────────────
                "changeStatus":                             "2",
                "dataOperation":                            "",
                "concurrencyHash":                          "",
                "disallowedActions":                        "",
                "disallowedActionsMessage":                 "",
                "isPredefaulted":                           False,
                "isDomainRestricted":                       False,
                "lastModifiedDate":                         "",
                "lastModifiedTime":                         0,
                "lastModifiedUser":                         "",

                # ── Active / Type ─────────────────────────────────────────────
                "isActive":                                 bool_val("Active"),
                "supplierTypeCode":                         val("Supplier Type"),
                "supplierTypeID":                           0,
                "supplierTypeDescription":                  "",
                "purchaseTypeCode":                         val("Purchase Type"),
                "purchaseTypeID":                           0,
                "purchaseTypeDescription":                  "",
                "purchaseCodeID":                           0,

                # ── Currency / Language ───────────────────────────────────────
                "currencyCode":                             val("Currency"),
                "currencyDescription":                      "",
                "currencyID":                               0,
                "languageCode":                             val("Language"),
                "languageDescription":                      "",

                # ── Credit Terms ──────────────────────────────────────────────
                "creditTermsCode":                          val("Credit Terms"),
                "creditTermsID":                            0,
                "creditTermsDescription":                   "",
                "normalCreditTermsID":                      0,
                "normalCreditTermsType":                    "",

                # ── Invoice / Status ──────────────────────────────────────────
                "invoiceStatusCode":                        val("Invoice Status"),
                "invoiceStatusDescription":                 "",
                "invoiceStatusID":                          0,
                "isIndividualPayments":                     False,
                "isSplitAccount":                           False,

                # ── GL Profiles ───────────────────────────────────────────────
                "invoiceControlGLProfileCode":              val("Invoice Control GL Profile"),
                "invoiceControlGLProfileDesc":              "",
                "invoiceControlGLProfileID":                0,
                "creditNoteControlGLProfileCode":           val("Credit Note Control GL Profile"),
                "creditNoteControlGLProfileDesc":           "",
                "creditNoteControlGLProfileID":             0,
                "prePaymentControlGLProfileCode":           val("Prepayment Control GL Profile"),
                "prePaymentControlGLProfileDesc":           "",
                "prePaymentControlGLProfileID":             0,
                "purchaseAccountGLProfileCode":             val("Purchase Account GL Profile"),
                "purchaseAccountProfileDesc":               "",
                "purchaseAccountGLProfileID":               0,
                "financeChargeGLProfileID":                 0,

                # ── Tax ───────────────────────────────────────────────────────
                "taxZone":                                  val("Tax Zone"),
                "taxZoneDescription":                       "",
                "taxClass":                                 val("Tax Class" or ""),
                "taxClassDescription":                      "",
                "taxDeclaration":                           0,
                "taxUsage":                                 val("Tax Usuage"),   # note: matches Excel typo
                "taxUsageDescription":                      "",
                "taxLevel":                                 "",
                "taxNature":                                "SERVICE",
                "taxIDFiscalCode":                          "",
                "foreignFiscalCode":                        "",
                "chamberOfCommerceNumber":                  "",
                "birthCity":                                "",
                "TIDNotice":                                "",
                "whtCertFormatCode":                        "",
                "whtCertFormatDescription":                 "",
                "whtCertFormatID":                          0,
                "isTaxableSupplier":                        bool_val("Taxable (Yes/No)"),
                "isTaxInCity":                              bool_val("Tax in City(Yes/No)"),
                "isTaxIncluded":                            bool_val("Tax Included(Yes/No)"),
                "isTaxReport":                              bool_val("Tax Report"),
                "isTaxReportForBusinessRelation":           False,
                "isTaxConfirmed":                           False,
                "isWithholdingTax":                         False,
                "isLastFiling":                             False,
                "isReportedIN":                             False,
                "federalTax":                               val("Federal tax"),
                "stateTax":                                 val("State Tax"),
                "miscellaneousTax1":                        val("Miscellaneous Tax 1"),
                "miscellaneousTax2":                        val("Miscellaneous Tax 2"),
                "miscellaneousTax3":                        val("Miscellaneous Tax 3"),
                "nameControl":                              "",
                "EORINumber":                               "",

                # ── Payment ───────────────────────────────────────────────────
                "paymentGroupCode":                         "",
                "paymentGroupDescription":                  "",
                "paymentGroupID":                           0,
                "isPayBankCharge":                          False,
                "isCompensationAllowed":                    False,
                "externalCustomerNumber":                   "",
                "deliveryConditionID":                      0,

                # ── Remittance ────────────────────────────────────────────────
                "isRemittanceRequired":                     False,
                "isSendRemittance":                         False,
                "remittanceAddressChangeStatus":            "",
                "remittanceAddressTypeCode":                "",
                "remitSupplierContactV2s":                  [],

                # ── Corporate Group ───────────────────────────────────────────
                "corporateGroupCode":                       "",
                "corporateGroupDescription":                "",
                "corporateGroupID":                         0,

                # ── Sub Account ───────────────────────────────────────────────
                "subAccountProfileCode":                    "",
                "subAccountProfileDesc":                    "",
                "subAccountProfileID":                      0,

                # ── Contact ───────────────────────────────────────────────────
                "EMail":                                    val("Email"),
                "fax":                                      "",
                "telephone":                                val("Telephone"),
                "webSite":                                  "",
                "commentNote":                              "",
                "creditAgencyReference":                    "",

                # ── Custom Fields ─────────────────────────────────────────────
                "customShort0":  "", "customShort1":  "", "customShort2":  "",
                "customShort3":  "", "customShort4":  "", "customShort5":  "",
                "customShort6":  "", "customShort7":  "", "customShort8":  "",
                "customShort9":  "", "customShort10": "", "customShort11": "",
                "customShort12": "", "customShort13": "", "customShort14": "",
                "customShort15": "", "customShort16": "", "customShort17": "",
                "customShort18": "", "customShort19": "",
                "customLong0":   "", "customLong1":   "",
                "customNote":    "",
                "customCombo0":  "", "customCombo1":  "", "customCombo2":  "",
                "customCombo3":  "", "customCombo4":  "", "customCombo5":  "",
                "customCombo6":  "", "customCombo7":  "", "customCombo8":  "",
                "customCombo9":  "", "customCombo10": "", "customCombo11": "",
                "customCombo12": "", "customCombo13": "", "customCombo14": "",
                "customDecimal0": 0, "customDecimal1": 0, "customDecimal2": 0,
                "customDecimal3": 0, "customDecimal4": 0,
                "customInteger0": 0, "customInteger1": 0, "customInteger2": 0,
                "customInteger3": 0, "customInteger4": 0,

                # ── Sub-lists ─────────────────────────────────────────────────
                "supplierContactV2s":                       [],
                "supplierSafDefaultV2s":                    [],
                "supplierVatV2s":                           [],
                "bankNumberRefV2s":                         [],
                "MDMBankNrSharedSets":                      [],
            }
        ],
    }


# =============================================================================
# 3. SUPPLIER API
# =============================================================================

class _TokenExpired(Exception):
    pass


def post_supplier(payload: dict, token: str, is_create: bool = False) -> tuple[bool, str]:
    """
    POST supplier payload to QAD.
    Returns (success, error_msg). Raises _TokenExpired on 401.

    CREATE → viewUri only  (no sharedSetCode/supplierCode in query string)
    UPDATE → sharedSetCode + supplierCode + viewUri
    """
    supplier      = payload["supplierV2s"][0]
    shared_set    = supplier.get("sharedSetCode", "")
    supplier_code = supplier.get("supplierCode", "")

    if is_create:
        # QAD UI sends only viewUri for new records — adding supplierCode causes
        # a record-lookup that fails because the record doesn't exist yet
        url = (
            f"{CONFIG['qad']['base_url']}/api/erp/supplierV2s"
            f"?viewUri=urn:be:com.qad.base.supplier.ISupplierV2"
        )
    else:
        url = (
            f"{CONFIG['qad']['base_url']}/api/erp/supplierV2s"
            f"?sharedSetCode={shared_set}&supplierCode={supplier_code}"
            f"&viewUri=urn:be:com.qad.base.supplier.ISupplierV2"
        )

    resp = requests.post(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/json",
        },
        json=payload,
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    resp_json = resp.json()
    submit    = resp_json.get("submitResult", {})

    if submit.get("success") is True:
        return True, ""

    errors    = submit.get("errors", [])

    # ── DEBUG: print exact field names QAD rejected ───────────────────────
    import json
    print(f"\nDEBUG post_supplier full response:")
    print(json.dumps(resp_json.get("submitResult", {}), indent=2))
    #---- DEBUG END ------------------------------------


    error_msg = "; ".join(
        e.get("message", "").strip()
        for e in errors
        if e.get("message", "").strip()
    ) or f"HTTP {resp.status_code} — submitResult.success was not True"

    return False, error_msg


def get_supplier(shared_set: str, supplier_code: str, token: str) -> dict:
    url = f"{CONFIG['qad']['base_url']}/api/erp/supplierV2s"

    resp = requests.get(
        url,
        headers={
            "Authorization": f"Bearer {token}",
        },
        params={
            "sharedSetCode": shared_set,
            "supplierCode":  supplier_code,
            "viewUri":       "urn:be:com.qad.base.supplier.ISupplierV2",
        },
        timeout=30,
    )

    return resp.json()


# =============================================================================
# 4. DOMAIN SETTINGS API  (mfgSuppliers)
# =============================================================================

def get_mfg_supplier(domain: str, supplier_code: str, token: str) -> dict:
    """Fetch the auto-created mfgSupplier domain record."""
    url = f"{CONFIG['qad']['base_url']}/api/erp/mfgSuppliers"

    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params={
            "domainContext": domain,
            "supplierCode":  supplier_code,
            "viewUri":       "urn:be:com.qad.base.supplier.IMfgSupplier",
        },
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    return resp.json()


def post_mfg_supplier(payload: dict, domain: str, supplier_code: str, token: str) -> tuple[bool, str]:
    """POST updated domain settings back to QAD."""
    url = (
        f"{CONFIG['qad']['base_url']}/api/erp/mfgSuppliers"
        f"?domainContext={domain}&supplierCode={supplier_code}"
        f"&viewUri=urn:be:com.qad.base.supplier.IMfgSupplier"
    )

    resp = requests.post(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/json",
        },
        json=payload,
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    resp_json = resp.json()

    # ── DEBUG BLOCK (remove before production) ────────────────────────────
    #print("\n" + "="*60)
    #print("DEBUG: POST mfgSuppliers response")
    #print("="*60)
    #print(f"  HTTP Status   : {resp.status_code}")
    #print(f"  Supplier      : {supplier_code}")
    #print(f"  Domain        : {domain}")
    #print(f"  Submit Success: {resp_json.get('submitResult', {}).get('success')}")
    #errors = resp_json.get("submitResult", {}).get("errors", [])
    #print(f"  Errors ({len(errors)}):")
    #for e in errors:
    #    print(f"    - field   : {e.get('field', 'N/A')}")
    #    print(f"      message : {e.get('message', 'N/A')}")
    #    print(f"      value   : {e.get('value', 'N/A')}")
    #    print(f"      code    : {e.get('code', 'N/A')}")
    #print("\n  Full submitResult:")
    #print(json.dumps(resp_json.get("submitResult", {}), indent=4))
    #print("="*60 + "\n")
    # ── END DEBUG BLOCK ───────────────────────────────────────────────────

    submit = resp_json.get("submitResult", {})

    if submit.get("success") is True:
        return True, ""

    errors    = submit.get("errors", [])
    error_msg = "; ".join(
        e.get("message", "").strip()
        for e in errors
        if e.get("message", "").strip()
    ) or f"HTTP {resp.status_code} — submitResult.success was not True"

    return False, error_msg


def update_domain_settings(
    supplier_code: str,
    domain:        str,
    site_code:     str,
    daybook_set:   str,
    token:         str,
) -> tuple[bool, str]:
    """
    GET auto-created mfgSupplier → patch siteCode + daybookSetCode → POST back.
    Returns (success, error_msg).
    """
    existing = get_mfg_supplier(domain, supplier_code, token)

    mfg_list = (
        existing.get("data", {}).get("mfgSuppliers")
        or existing.get("mfgSuppliers")
    )

    if not mfg_list:
        return False, f"Domain settings record not found for supplier '{supplier_code}' in domain '{domain}'"

    payload    = existing.get("data", existing)
    mfg_record = mfg_list[0]

    mfg_record["siteCode"]       = site_code
    mfg_record["daybookSetCode"] = daybook_set

    return post_mfg_supplier(payload, domain, supplier_code, token)


# =============================================================================
# 5. WORKBOOK HELPERS
# =============================================================================

def _mark_done(ws, row_idx: int, status_col: int, error_col: int):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col,  value="")


def _mark_error(
    ws,
    row_idx:    int,
    status_col: int,
    error_col:  int,
    header_row: list,
    bad_cols:   list,
    error_msg:  str,
):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL

    ws.cell(row=row_idx, column=1).fill          = RED_FILL
    ws.cell(row=row_idx, column=status_col).fill = RED_FILL
    ws.cell(row=row_idx, column=error_col).fill  = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="ERROR")
    ws.cell(row=row_idx, column=error_col,  value=error_msg)

    for col_name in bad_cols:
        if col_name in header_row:
            ws.cell(row=row_idx, column=header_row.index(col_name) + 1).fill = RED_FILL


# =============================================================================
# 6. MANDATORY FIELD CHECK
# =============================================================================

def _check_mandatory(row_data: dict) -> list[str]:
    missing = []
    for col in MANDATORY_COLUMNS:
        v = row_data.get(col)
        if v is None or str(v).strip() == "" or str(v).strip().lower() == "none":
            missing.append(col)
    return missing


def _check_mandatory_domain(row_data: dict) -> list[str]:
    missing = []
    for col in MANDATORY_DOMAIN_COLUMNS:
        v = row_data.get(col)
        if v is None or str(v).strip() == "" or str(v).strip().lower() == "none":
            missing.append(col)
    return missing


# =============================================================================
# 7. SINGLE-FILE PROCESSOR
# =============================================================================

def process_file(file_path: str, tm: TokenManager) -> tuple[int, int]:
    """Open one workbook, process every row, return (ok_count, fail_count)."""

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Row 1 = section labels, Row 2 = column headers, Row 3+ = data
    header_row = [
        str(c.value).strip() if c.value is not None else ""
        for c in ws[2]
    ]

    for col_name in ("Status", "Error"):
        if col_name not in header_row:
            ws.cell(row=2, column=len(header_row) + 1, value=col_name)
            header_row.append(col_name)

    status_col = header_row.index("Status") + 1
    error_col  = header_row.index("Error")  + 1

    ok_count   = 0
    fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3), start=3):

        row_values = [cell.value for cell in row]

        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        status         = str(row_data.get("Status", "")).strip().upper()
        data_operation = str(row_data.get("Data Operation", "")).strip().upper() or "C"

        # Skip completed CREATE rows only
        if status == "DONE" and data_operation == "C":
            continue

        # ── Mandatory check ───────────────────────────────────────────────
        missing = _check_mandatory(row_data)
        if missing:
            fail_count += 1
            _mark_error(
                ws, row_idx, status_col, error_col, header_row,
                missing,
                f"Missing mandatory fields: {', '.join(missing)}",
            )
            wb.save(file_path)
            continue

        # ── Operation validity check ──────────────────────────────────────
        if data_operation not in {"C", "U"}:
            fail_count += 1
            _mark_error(
                ws, row_idx, status_col, error_col, header_row,
                ["Data Operation"],
                "Data Operation must be C, U, or blank",
            )
            wb.save(file_path)
            continue

        # ── Domain settings mandatory check (CREATE only) ─────────────────
        is_create = (data_operation == "C")

        if is_create:
            missing_domain = _check_mandatory_domain(row_data)
            if missing_domain:
                fail_count += 1
                _mark_error(
                    ws, row_idx, status_col, error_col, header_row,
                    missing_domain,
                    f"Missing domain setting fields: {', '.join(missing_domain)}",
                )
                wb.save(file_path)
                continue

        # ── Build payload ─────────────────────────────────────────────────
        if data_operation == "U":
            # GET existing object → hydrate → patch editable fields → POST back
            existing = get_supplier(
                row_data.get("Shared Set", ""),
                row_data.get("Supplier", ""),
                tm.get(),
            )

            if not existing.get("data") or not existing["data"].get("supplierV2s"):
                fail_count += 1
                _mark_error(
                    ws, row_idx, status_col, error_col, header_row, [],
                    f"UPDATE failed: Supplier '{row_data.get('Supplier', '')}' not found in QAD",
                )
                wb.save(file_path)
                continue

            payload  = existing["data"]
            supplier = payload["supplierV2s"][0]

            supplier["businessRelationCode"]          = str(row_data.get("Business Relation", "")).strip()
            supplier["isActive"]                      = str(row_data.get("Active", "")).strip().lower() == "yes"
            supplier["supplierTypeCode"]              = str(row_data.get("Supplier Type")).strip()
            supplier["purchaseTypeCode"]              = str(row_data.get("Purchase Type")).strip()
            supplier["addressSearchName"]             = str(row_data.get("Search Name")).strip()
            supplier["city"]                          = str(row_data.get("City")).strip()
            supplier["stateCode"]                     = str(row_data.get("State")).strip()
            supplier["street1"]                       = str(row_data.get("Address 1")).strip()
            supplier["street2"]                       = str(row_data.get("Address 2")).strip()
            supplier["zipCode"]                       = str(row_data.get("Postal Code")).strip()
            supplier["telephone"]                     = str(row_data.get("Telephone")).strip()
            supplier["EMail"]                         = str(row_data.get("Email")).strip()
            supplier["currencyCode"]                  = str(row_data.get("Currency", "")).strip()
            supplier["creditTermsCode"]               = str(row_data.get("Credit Terms", "")).strip()
            supplier["invoiceStatusCode"]             = str(row_data.get("Invoice Status", "")).strip()
            supplier["invoiceControlGLProfileCode"]   = str(row_data.get("Invoice Control GL Profile", "")).strip()
            supplier["creditNoteControlGLProfileCode"]= str(row_data.get("Credit Note Control GL Profile", "")).strip()
            supplier["prePaymentControlGLProfileCode"]= str(row_data.get("Prepayment Control GL Profile", "")).strip()
            supplier["purchaseAccountGLProfileCode"]  = str(row_data.get("Purchase Account GL Profile", "")).strip()
            supplier["taxZone"]                       = str(row_data.get("Tax Zone", "")).strip()

        else:
            payload = build_payload(row_data)

        # ── Step 1: POST supplier (with one token-refresh retry on 401) ────
        success   = False
        error_msg = ""

        for attempt in range(2):
            try:
                success, error_msg = post_supplier(payload, tm.get(), is_create=is_create)
                break
            except _TokenExpired:
                if attempt == 0:
                    tm.refresh()
                    continue
                error_msg = "Token refresh failed — unauthorised"
                break
            except requests.RequestException as e:
                error_msg = f"Network error: {e}"
                break

        if not success:
            fail_count += 1
            _mark_error(ws, row_idx, status_col, error_col, header_row, [], error_msg)
            wb.save(file_path)
            time.sleep(0.1)
            continue

        # ── Step 2 (CREATE only): update domain settings ──────────────────
        if is_create:
            domain        = str(row_data.get("Domain", "")).strip()
            site_code     = str(row_data.get("Site Code", "")).strip()
            daybook_set   = str(row_data.get("Daybook Set", "")).strip()
            supplier_code = str(row_data.get("Supplier", "")).strip()

            domain_ok  = False
            domain_err = ""

            for attempt in range(2):
                try:
                    domain_ok, domain_err = update_domain_settings(
                        supplier_code, domain, site_code, daybook_set, tm.get()
                    )
                    break
                except _TokenExpired:
                    if attempt == 0:
                        tm.refresh()
                        continue
                    domain_err = "Token refresh failed — unauthorised"
                    break
                except requests.RequestException as e:
                    domain_err = f"Network error: {e}"
                    break

            if not domain_ok:
                fail_count += 1
                _mark_error(
                    ws, row_idx, status_col, error_col, header_row, [],
                    f"Supplier created but domain settings failed: {domain_err}",
                )
                wb.save(file_path)
                time.sleep(0.1)
                continue

        # ── All steps passed ──────────────────────────────────────────────
        ok_count += 1
        _mark_done(ws, row_idx, status_col, error_col)
        wb.save(file_path)
        time.sleep(0.1)

    return ok_count, fail_count


# =============================================================================
# 8. ORCHESTRATOR  (folder scan → process every file → return totals)
# =============================================================================

def run(folder_path: str) -> tuple[int, int]:
    """
    Scan folder_path for .xlsx files, process each one, return (total_ok, total_fail).
    This is the single entry point used by both main.py and __main__.
    """
    folder = os.path.abspath(folder_path)

    if not os.path.exists(folder):
        print(f"ERROR: Folder not found: {folder}")
        raise RuntimeError(f"Folder not found: {folder}")

    xlsx_files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not xlsx_files:
        print(f"ERROR: No .xlsx files found in: {folder}")
        raise RuntimeError(f"No .xlsx files found in: {folder}")

    tm = TokenManager()

    total_ok   = 0
    total_fail = 0

    for file_name in xlsx_files:
        file_path = os.path.join(folder, file_name)
        print(f"Loading : {file_path}")

        ok, fail   = process_file(file_path, tm)
        total_ok   += ok
        total_fail += fail

        print("\nLoad Summary")
        print("-" * 40)
        print(f"  Rows loaded successfully : {ok}")
        print(f"  Rows failed              : {fail}")
        if fail == 0:
            print("  Result : ALL ROWS LOADED SUCCESSFULLY ✓")
        else:
            print("  Result : COMPLETED WITH ERRORS — fix red rows and re-run")
        print()

    print("=" * 40)
    print(f"  Total loaded : {total_ok}")
    print(f"  Total failed : {total_fail}")
    print("=" * 40)

    return total_ok, total_fail


# =============================================================================
# 9. ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    folder = os.path.abspath(
        os.path.join(ROOT_DIR, CONFIG["folders"]["supplier"])
    )
    ok, fail = run(folder)
    sys.exit(0 if fail == 0 else 1)