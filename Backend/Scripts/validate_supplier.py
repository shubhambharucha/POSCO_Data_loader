import openpyxl
from openpyxl.styles import PatternFill

# ---------------------------------------------------------------------------
# Fill constants
# ---------------------------------------------------------------------------
RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

# ---------------------------------------------------------------------------
# Field rules
# ---------------------------------------------------------------------------
FIELD_RULES = {
    "Supplier":            {"max_len": 8,    "type": "character"},
    "Shared Set":          {"max_len": 20,   "type": "character"},
    "Business Relation":   {"max_len": 20,   "type": "character"},
    "Active":              {"max_len": None, "type": "logical"},
    "Currency":            {"max_len": 3,    "type": "character"},
    "Credit Terms":        {"max_len": 8,    "type": "character"},
    "Invoice Status":      {"max_len": 20,   "type": "character"},
    "Invoice Control GL Profile": {"max_len": 20,   "type": "character"},
    "Credit Note Control GL Profile": {"max_len": 20,   "type": "character"},
    "Prepayment Control GL Profile": {"max_len": 20,   "type": "character"},
    "Purchase Account GL Profile": {"max_len": 20,   "type": "character"},
}

ENTITY_COL    = "Supplier"
VALID_LOGICAL = {"yes", "no"}

# Rows already processed by validation/load
SKIP_STATUSES = {"DONE", "READY"}


def validate(file_path):
    """
    Validate a supplier Excel template.

    Behaviour:
    - Strips whitespace from headers
    - Skips READY/DONE rows
    - Skips empty rows
    - Writes validation errors into Error column
    - Marks good rows as READY
    - Highlights invalid cells in RED
    - Saves workbook in place

    Returns:
        dict containing validation summary
    """

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # -------------------------------------------------------------------
    # Read and normalize headers
    # -------------------------------------------------------------------
    raw_headers = [cell.value for cell in ws[1]]
    header_row  = [str(h).strip() if h is not None else "" for h in raw_headers]

    # -------------------------------------------------------------------
    # Ensure operational columns exist
    # -------------------------------------------------------------------
    if "Status" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Status")
        header_row.append("Status")

    if "Error" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Error")
        header_row.append("Error")

    # -------------------------------------------------------------------
    # Required entity column check
    # -------------------------------------------------------------------
    if ENTITY_COL not in header_row:
        raise ValueError(f"Required column missing: {ENTITY_COL}")

    # -------------------------------------------------------------------
    # Column indexes
    # -------------------------------------------------------------------
    status_col_idx = header_row.index("Status") + 1
    error_col_idx  = header_row.index("Error") + 1
    entity_col_idx = header_row.index(ENTITY_COL) + 1

    # -------------------------------------------------------------------
    # Validation stats
    # -------------------------------------------------------------------
    has_errors      = False
    rows_processed  = 0
    rows_passed     = 0
    rows_failed     = 0
    rows_skipped    = 0

    # -------------------------------------------------------------------
    # Process rows
    # -------------------------------------------------------------------
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):

        row_values = [cell.value for cell in row]

        # ---------------------------------------------------------------
        # Skip fully empty rows
        # ---------------------------------------------------------------
        if not any(row_values):
            rows_skipped += 1
            continue

        row_data = dict(zip(header_row, row_values))

        # ---------------------------------------------------------------
        # Skip READY / DONE rows
        # ---------------------------------------------------------------
        current_status = str(row_data.get("Status", "") or "").strip().upper()

        if current_status in SKIP_STATUSES:
            rows_skipped += 1
            continue

        rows_processed += 1

        row_errors      = []
        error_cell_idxs = []

        # ---------------------------------------------------------------
        # Field-level validation
        # ---------------------------------------------------------------
        for col_name, rule in FIELD_RULES.items():

            if col_name not in header_row:
                continue

            col_idx = header_row.index(col_name) + 1

            value    = row_data.get(col_name)
            str_val  = str(value).strip() if value is not None else ""
            is_empty = str_val == "" or str_val.lower() == "none"

            # -----------------------------------------------------------
            # Empty check
            # -----------------------------------------------------------
            if is_empty:
                row_errors.append(f"{col_name}: empty")
                error_cell_idxs.append(col_idx)
                continue

            # -----------------------------------------------------------
            # Logical validation
            # -----------------------------------------------------------
            if rule["type"] == "logical":

                if str_val.lower() not in VALID_LOGICAL:
                    row_errors.append(
                        f"{col_name}: must be Yes or No (got '{str_val}')"
                    )
                    error_cell_idxs.append(col_idx)

                continue

            # -----------------------------------------------------------
            # Character length validation
            # -----------------------------------------------------------
            max_len = rule.get("max_len")

            if max_len and len(str_val) > max_len:
                row_errors.append(
                    f"{col_name}: max {max_len} chars (got {len(str_val)})"
                )
                error_cell_idxs.append(col_idx)

        # ---------------------------------------------------------------
        # Clear old formatting
        # ---------------------------------------------------------------
        for cell in ws[row_idx]:
            cell.fill = CLEAR_FILL

        # ---------------------------------------------------------------
        # Write validation results
        # ---------------------------------------------------------------
        if row_errors:

            has_errors = True
            rows_failed += 1

            error_msg = "; ".join(row_errors)

            # Highlight entity cell
            ws.cell(row=row_idx, column=entity_col_idx).fill = RED_FILL

            # Highlight invalid cells
            for cidx in error_cell_idxs:
                ws.cell(row=row_idx, column=cidx).fill = RED_FILL

            # Set Error column
            ws.cell(
                row=row_idx,
                column=error_col_idx,
                value=error_msg
            ).fill = RED_FILL

            # Clear status
            ws.cell(
                row=row_idx,
                column=status_col_idx,
                value="ERROR"
            )

        else:

            rows_passed += 1

            # Row is valid
            ws.cell(
                row=row_idx,
                column=status_col_idx,
                value="READY"
            )

            ws.cell(
                row=row_idx,
                column=error_col_idx,
                value=""
            )

    # -------------------------------------------------------------------
    # Save workbook
    # -------------------------------------------------------------------
    wb.save(file_path)

    # -------------------------------------------------------------------
    # Return validation summary
    # -------------------------------------------------------------------
    return {
        "has_errors": has_errors,
        "rows_processed": rows_processed,
        "rows_passed": rows_passed,
        "rows_failed": rows_failed,
        "rows_skipped": rows_skipped,
    }


# ---------------------------------------------------------------------------
# Standalone testing mode
# ---------------------------------------------------------------------------
def run_standalone():

    import sys
    import os

    #Temporary fix for standalone execution in cmd
    ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    
    sys.path.append(ROOT_DIR)

    from config import CONFIG

    folder = os.path.abspath(os.path.join(ROOT_DIR, CONFIG["folders"]["supplier"]))

    if not os.path.exists(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)
    
    xlsx_files = [f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")]

    if not xlsx_files:
        print(f"ERROR: No .xlsx file found in folder: {folder}")
        sys.exit(1)

    total_processed = 0
    total_failed   = 0

    for file_name in xlsx_files:

        file_path = os.path.join(folder, file_name)

        print(f"Validating: {file_path} ...")

        result = validate(file_path)

        print("\nValidation Summary")
        print("------------------------------")
        print(f"Rows processed : {result['rows_processed']}")
        print(f"Rows passed    : {result['rows_passed']}")
        print(f"Rows failed    : {result['rows_failed']}")
        print(f"Rows skipped   : {result['rows_skipped']}")

        total_processed += result["rows_processed"]
        total_failed    += result["rows_failed"]

        if result["has_errors"]:
            print("\nValidation FAILED")
        else:
            print("\nValidation PASSED — all rows are READY for Loading.")

    #outside loop    
    sys.exit(0 if total_failed == 0 else 1)

if __name__ == "__main__":
    run_standalone()