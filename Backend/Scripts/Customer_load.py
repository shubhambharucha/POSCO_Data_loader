"""
Customer_load.py
----------------
Loads customer rows from all .xlsx files in the configured customer folder
into QAD via the customerV2s API.

Behaviour
---------
- Fetches one OAuth token per run; refreshes only on 401
- Skips rows with Status = DONE (for Create operations only)
- Lightweight mandatory-field check before any API call
- Success check: submitResult.success == True
- Error messages extracted from submitResult.errors[].message
- On success  → clear all red fills, clear Error, set Status = DONE
- On failure  → red-fill first col + bad cols + Error col, log message, set Status = ERROR
- Processes every row regardless of individual failures
- Returns (ok_count, fail_count)

CREATE flow (Data Operation = C):
  1. POST to customerV2s  → creates customer
  2. GET  mfgCustomers    → fetch auto-created domain record
  3. PATCH + POST         → update siteCode + daybookSetCode
  If step 3 fails → row marked ERROR with "Customer created but domain settings failed: ..."

UPDATE flow (Data Operation = U):
  1. GET  customerV2s     → fetch existing record
  2. Patch editable fields
  3. POST customerV2s     → update customer
  (Domain settings NOT touched on UPDATE — one-time setup only)
"""

import os
import sys
import time
import json
import requests
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

# ── Path / config ─────────────────────────────────────────────────────────
ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT_DIR)
from config import CONFIG

# ── Fill constants ────────────────────────────────────────────────────────
RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

# ── Mandatory columns (customer) ──────────────────────────────────────────
MANDATORY_COLUMNS = [
    "Customer",
    "Shared Set",
    "Business Relation",
    "Active",
    "Currency",
    "Credit Terms",
    "Invoice Status",
    "Invoice Control GL Profile",
    "Credit Note Control GL Profile",
    "Prepayment Control GL Profile",
    "Sales Account GL Profile",
]

# ── Mandatory columns for domain settings (CREATE only) ───────────────────
MANDATORY_DOMAIN_COLUMNS = [
    "Domain",
    "Site Code",
    "Daybook Set",
]


# =============================================================================
# 1. AUTHENTICATION
# =============================================================================

def _fetch_token() -> str:
    url  = f"{CONFIG['qad']['base_url']}/oauth/token"
    resp = requests.post(url, data=CONFIG["qad"]["auth"], timeout=30)
    resp.raise_for_status()
    token = resp.json().get("access_token")
    if not token:
        raise RuntimeError("OAuth response did not contain access_token")
    return token


class TokenManager:
    """Holds one token for the run; refreshes on demand (401)."""

    def __init__(self):
        self._token: str | None = None

    def get(self) -> str:
        if self._token is None:
            self._token = _fetch_token()
        return self._token

    def refresh(self) -> str:
        self._token = _fetch_token()
        return self._token


# =============================================================================
# 2. PAYLOAD BUILDER  (full payload — every field visible)
# =============================================================================

def build_payload(row: dict) -> dict:
    def val(col: str) -> str:
        v = row.get(col)
        return str(v).strip() if v is not None else ""

    def bool_val(col: str) -> bool:
        return val(col).lower() == "yes"

    def float_val(col: str) -> float:
        return float(val(col) or 0)

    def int_val(col: str) -> int:
        return int(float(val(col) or 0))

    def date_val(col: str) -> str:
        v = val(col)
        if not v:
            return ""
        try:
            #adjust input format based on source (excel date format)
            return datetime.strtime(v, "%d-%m-%Y").strftime("%Y-%m-%d")
        except:
            return ""

    customer_code = val("Customer")
    shared_set    = val("Shared Set")
    uri           = f"urn:be:com.qad.base.customer.ICustomerV2:{shared_set}.{customer_code}"

    return {
        "supplementaryMessages": [],
        "customerV2s": [
            {
                # ── URIs / Identity ───────────────────────────────────────────
                "uri":                                      uri,
                "instanceURI":                              uri,
                "customerCode":                             customer_code,
                "customerID":                               0,
                "sharedSetCode":                            shared_set,
                "sharedSetID":                              0,

                # ── Business Relation ─────────────────────────────────────────
                "businessRelationCode":                     val("Business Relation"),
                "businessRelationID":                       0,
                "businessRelationName":                     "",
                "businessRelationName2":                    "",
                "businessRelationName3":                    "",
                "businessRelationConcurrencyHash":          "",
                "isBusinessRelationActive":                 True,
                "isBusinessRelationFieldsEnabled":          True,
                "isBusinessRelationIntercompany":           False,
                "isCreateBusinessRelationRequired":         True,
                "isInternalEntity":                         False,
                "intercompanyCode":                         "",

                # ── Address ───────────────────────────────────────────────────
                "addressID":                                0,
                "addressName":                              val("Name"),
                "addressSearchName":                        val("Search Name"),
                "addressTypeCode":                          val("Address Type"),
                "addressConcurrencyHash":                   "",
                "street1":                                  val("Address 1"),
                "street2":                                  val("Address 2"),
                "street3":                                  "",
                "city":                                     val("City"),
                "cityCode":                                 "",
                "zipCode":                                  val("Postal Code"),
                "stateCode":                                val("State"),
                "stateDescription":                         "",
                "stateTax":                                 val("State Tax"),
                "countryCode":                              val("Country"),
                "countryDescription":                       "",
                "countyCode":                               "",
                "countyDescription":                        "",
                "postalFormat":                             "0",
                "latitude":                                 0,
                "longitude":                                0,
                "isTemporaryAddress":                       False,

                # ── Status / Concurrency ──────────────────────────────────────
                "changeStatus":                             "2",
                "dataOperation":                            "",
                "concurrencyHash":                          "",
                "disallowedActions":                        "",
                "disallowedActionsMessage":                 "",
                "isPredefaulted":                           False,
                "isDomainRestricted":                       False,
                "lastModifiedDate":                         "",
                "lastModifiedTime":                         0,
                "lastModifiedUser":                         "",

                # ── Active / Type ─────────────────────────────────────────────
                "isActive":                                 bool_val("Active"),
                "customerTypeCode":                         val("Customer Type"),
                "customerTypeID":                           0,

                # ── Currency / Language ───────────────────────────────────────
                "currencyCode":                             val("Currency"),
                "currencyDescription":                      "",
                "currencyID":                               0,
                "customerCurrencyCode":                     val("Currency"),
                "languageCode":                             val("Language"),
                "languageDescription":                      "",

                # ── Credit Terms ──────────────────────────────────────────────
                "creditTermsCode":                          val("Credit Terms"),
                "creditTermsDescription":                   "",
                "creditTermsID":                            0,
                "creditTermsType":                          "",

                # ── Invoice / Status ──────────────────────────────────────────
                "invoiceStatusCode":                        val("Invoice Status"),
                "invoiceStatusDescription":                 "",
                "invoiceStatusID":                          0,
                "isInvoiceByAuthorization":                 False,
                "isWithPreInvoiceGroup":                    False,
                "isPrintBillWithItemDetails":               False,

                # ── GL Profiles ───────────────────────────────────────────────
                "invoiceControlGLProfileCode":              val("Invoice Control GL Profile"),
                "invoiceControlGLProfileDesc":              "",
                "invoiceControlGLProfileID":                0,
                "creditNoteControlGLProfileCode":           val("Credit Note Control GL Profile"),
                "creditNoteControlGLProfileDesc":           "",
                "creditNoteControlGLProfileID":             0,
                "prePaymentControlGLProfileCode":           val("Prepayment Control GL Profile"),
                "prePaymentControlGLProfileDesc":           "",
                "prePaymentControlGLProfileID":             0,
                "salesAccountGLProfileCode":                val("Sales Account GL Profile"),
                "salesAccountGLProfileDesc":                "",
                "salesAccountGLProfileID":                  0,
                "deductionControlGLProfileCode":            "",
                "deductionControlGLProfileDesc":            "",
                "deductionControlGLProfileID":              0,
                "financeChargeGLProfileCode":               "",
                "financeChargeGLProfileDesc":               "",
                "financeChargeGLProfileID":                 0,

                # ── Revenue Recognition GL Profiles ───────────────────────────
                "accruedRevRecGLProfileCode":               "",
                "accruedRevRecGLProfileDesc":               "",
                "accruedRevRecGLProfileID":                 0,
                "deferredRevRecGLProfileCode":              "",
                "deferredRevRecGLProfileDesc":              "",
                "deferredRevRecGLProfileID":                0,
                "COGSAccruedRevRecGLProfileCode":           "",
                "COGSAccruedRevRecGLProfileDesc":           "",
                "COGSAccruedRevRecGLProfileID":             0,
                "COGSDeferredRevRecGLProfileCode":          "",
                "COGSDeferredRevRecGLProfileDesc":          "",
                "COGSDeferredRevRecGLProfileID":            0,
                "COGSOffsetRevRecGLProfileCode":            "",
                "COGSOffsetRevRecGLProfileDesc":            "",
                "COGSOffsetRevRecGLProfileID":              0,
                "isAutoCreateRevRecContracts":              False,
                "revRecRuleID":                             0,
                "reviewRequiredForContracts":               "",

                # ── Tax ───────────────────────────────────────────────────────
                "taxZone":                                  val("Tax Zone"),
                "taxZoneDescription":                       "",
                "taxClass":                                 val("Tax Class"),
                "taxClassDescription":                      "",
                "taxDeclaration":                           0,
                "taxUsage":                                 val("Tax Usage"),
                "taxUsageDescription":                      "",
                "isTaxable":                                bool_val("Taxable(Yes/No)"),
                "isTaxInCity":                              bool_val("Tax in City(Yes/No)"),
                "isTaxIncluded":                            bool_val("Tax Included(Yes/No)"),
                "isTaxReport":                              bool_val("Tax Report"),
                "isLastFiling":                             False,
                "isReportedIN":                             False,
                "isElectronicInvoiceIN":                    False,
                "federalTax":                               val("Federal Tax"),
                "stateTax":                                 val("State Tax"),
                "miscellaneousTax1":                        val("Miscellaneous Tax 1"),
                "miscellaneousTax2":                        val("Miscellaneous Tax 2"),
                "miscellaneousTax3":                        val("Miscellaneous Tax 3"),
                "customerGTVatTransType":                   "",
                "vatDeliveryType":                          "",
                "vatPercentageLevel":                       "",
                "nameControl":                              "",
                "EORINumber":                               "",

                # ── Credit Limit ──────────────────────────────────────────────
                "fixedCreditLimit":                         float_val("Fixed Credit Limit"), #done
                "highCredit":                               0,
                "isFixedCreditLimit":                       bool_val("Apply Fixed Ceiling"), #done
                "isTurnOverCreditLimit":                    bool_val("Apply % of Turnover"), #done
                "isMaxDaysOverdueCreditLimit":              bool_val("Apply Maximum Days Overdue"), #done 
                "maxDaysCreditLimit":                       int_val("Maximum Days Overdue"), #done
                "turnoverCreditLimitPercent":               float_val("Percentage of Turnover"), #done
                "isLockedCreditLimit":                      bool_val("Credit Hold"), #done
                "isToBeLockedCreditLimit":                  False,
                "warningCreditLimitPercent":                float_val("Warning Ceiling %"), #done
                "creditAgencyReference":                    val("Credit Agency Ref"), #done
                "creditRatingCode":                         val("Credit Rating"),#done
                "creditRatingID":                           0,

                # ── Credit Check ──────────────────────────────────────────────
                "isOverruleAllowedSOCreditLimit":           bool_val("Overrule Allowed SO(Yes/No)"),
                "isCheckBeforeSOCreditLimit":               bool_val("Calculate before Order Entry(Yes/No)"),
                "isCheckAfterSOCreditLimit":                bool_val("Calculate after Order Entry(Yes/No)"),
                "isCheckBeforeInvoiceCreditLimit":          bool_val("Calculate before Invoice(Yes/No)"),
                "isCheckAfterInvoiceCreditLimit":           bool_val("Calculate after Invoice Entry(Yes/No)"),
                "isOverAllowedInvoiceCreditLimit":          bool_val("Overrule Allowed Invoice(Yes/No)"),
                "isIncludeDraftCreditLimit":                bool_val("Include Drafts(Yes/No)"),
                "isIncludeOpenItemsCreditLimit":            bool_val("Include Open Items(Yes/No)"),
                "isIncludeSOCheckCreditLimit":              bool_val("Include Sales Orders(Yes/No)"),
                "overToleranceAmount":                      0,
                "overTolerancePercent":                     0,
                "shortToleranceAmount":                     0,
                "shortTolerancePercent":                    0,
                "totalDaysLate":                            0,
                "totalNumberOfInvoices":                    0,

                # ── Finance Charges / Reminders / Statements ──────────────────
                "isFinanceCharge":                          False,
                "isPrintReminder":                          False,
                "isReminderRequired":                       False,
                "isPrintStatement":                         False,
                "reminderCountReset":                       False,
                "reminderAddressChangeStatus":              "",
                "reminderAddressID":                        0,
                "reminderAddressTypeCode":                  "",
                "reminderCustContactV2s":                   [],

                # ── Payment ───────────────────────────────────────────────────
                "paymentGroupCode":                         "",
                "paymentGroupDescription":                  "",
                "paymentGroupID":                           0,
                "domiciliationNumber":                      0,
                "isToleranceFromOwnBank":                   False,
                "isCompensationAllowed":                    False,

                # ── Billing ───────────────────────────────────────────────────
                "billToCustomerCode":                       "",
                "billToCustomerID":                         0,
                "billCollectorCode":                        "",
                "billCollectorID":                          0,
                "billingScheduleCode":                      "",
                "billingScheduleDescription":               "",
                "billingScheduleID":                        0,
                "statementCycle":                           "",
                "subAccountProfileCode":                    "",
                "subAccountProfileDesc":                    "",
                "subAccountProfileID":                      0,

                # ── Corporate Group ───────────────────────────────────────────
                "corporateGroupCode":                       "",
                "corporateGroupDescription":                "",
                "corporateGroupID":                         0,

                # ── Contact ───────────────────────────────────────────────────
                "EMail":                                    val("Email"),
                "fax":                                      "",
                "telephone":                                val("Telephone"),
                "webSite":                                  "",
                "commentNote":                              "",

                # ── Deduction ─────────────────────────────────────────────────
                "customerIsInclDeduction":                  bool_val("Include Deductions"),

                # ── Custom Fields ─────────────────────────────────────────────
                "customShort0":  "", "customShort1":  "", "customShort2":  "",
                "customShort3":  "", "customShort4":  "", "customShort5":  "",
                "customShort6":  "", "customShort7":  "", "customShort8":  "",
                "customShort9":  "", "customShort10": "", "customShort11": "",
                "customShort12": "", "customShort13": "", "customShort14": "",
                "customShort15": "", "customShort16": "", "customShort17": "",
                "customShort18": "", "customShort19": "",
                "customLong0":   "", "customLong1":   "",
                "customNote":    "",
                "customCombo0":  "", "customCombo1":  "", "customCombo2":  "",
                "customCombo3":  "", "customCombo4":  "", "customCombo5":  "",
                "customCombo6":  "", "customCombo7":  "", "customCombo8":  "",
                "customCombo9":  "", "customCombo10": "", "customCombo11": "",
                "customCombo12": "", "customCombo13": "", "customCombo14": "",
                "customDecimal0": 0, "customDecimal1": 0, "customDecimal2": 0,
                "customDecimal3": 0, "customDecimal4": 0,
                "customInteger0": 0, "customInteger1": 0, "customInteger2": 0,
                "customInteger3": 0, "customInteger4": 0,

                # ── Sub-lists ─────────────────────────────────────────────────
                "customerContactV2s":                       [],
                "customerSafDefaultV2s":                    [],
                "bankNumberRefV2s":                         [],
                "MDMBankNrSharedSets":                      [],
            }
        ],
    }


# =============================================================================
# 3. CUSTOMER API
# =============================================================================

class _TokenExpired(Exception):
    pass


def post_customer(payload: dict, token: str, is_create: bool = False) -> tuple[bool, str]:
    """
    POST customer payload to QAD.
    Returns (success, error_msg). Raises _TokenExpired on 401.
    """
    customer      = payload["customerV2s"][0]
    shared_set    = customer.get("sharedSetCode", "")
    customer_code = customer.get("customerCode", "")

    if is_create:
        url = (
            f"{CONFIG['qad']['base_url']}/api/erp/customerV2s"
            f"?viewUri=urn:be:com.qad.base.customer.ICustomerV2"
        )
    else:
        url = (
            f"{CONFIG['qad']['base_url']}/api/erp/customerV2s"
            f"?sharedSetCode={shared_set}&customerCode={customer_code}"
            f"&viewUri=urn:be:com.qad.base.customer.ICustomerV2"
        )

    resp = requests.post(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/json",
        },
        json=payload,
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    resp_json = resp.json()


    # ── DEBUG BLOCK (remove before production) ────────────────────────────
    #print("\n" + "="*60)
    #print("DEBUG: POST customerV2s response")
    #print("="*60)
    #print(f"  HTTP Status   : {resp.status_code}")
    #print(f"  Customer      : {customer_code}")
    #print(f"  Is Create     : {is_create}")
    #print(f"  Submit Success: {resp_json.get('submitResult', {}).get('success')}")
    #errors = resp_json.get("submitResult", {}).get("errors", [])
    #print(f"  Errors ({len(errors)}):")
    #for e in errors:
    #    print(f"    - field   : {e.get('field', 'N/A')}")
    #    print(f"      message : {e.get('message', 'N/A')}")
    #   print(f"      value   : {e.get('value', 'N/A')}")
    #    print(f"      code    : {e.get('code', 'N/A')}")
    #print("\n  Full submitResult:")
    #print(json.dumps(resp_json.get("submitResult", {}), indent=4))
    #print("="*60 + "\n")
    # ── END DEBUG BLOCK ───────────────────────────────────────────────────


    submit = resp_json.get("submitResult", {})

    if submit.get("success") is True:
        return True, ""

    errors    = submit.get("errors", [])
    error_msg = "; ".join(
        e.get("message", "").strip()
        for e in errors
        if e.get("message", "").strip()
    ) or f"HTTP {resp.status_code} — submitResult.success was not True"

    return False, error_msg


def get_customer(shared_set: str, customer_code: str, token: str) -> dict:
    url = f"{CONFIG['qad']['base_url']}/api/erp/customerV2s"

    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params={
            "sharedSetCode": shared_set,
            "customerCode":  customer_code,
            "viewUri":       "urn:be:com.qad.base.customer.ICustomerV2",
        },
        timeout=30,
    )

    return resp.json()


# =============================================================================
# 4. DOMAIN SETTINGS API  (mfgCustomers)
# =============================================================================

def get_mfg_customer(domain: str, customer_code: str, token: str) -> dict:
    """Fetch the auto-created mfgCustomer domain record."""
    url = f"{CONFIG['qad']['base_url']}/api/erp/mfgCustomers"

    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params={
            "domainContext": domain,
            "customerCode":  customer_code,
            "viewUri":       "urn:be:com.qad.base.customer.IMfgCustomer",
        },
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    return resp.json()


def post_mfg_customer(payload: dict, domain: str, customer_code: str, token: str) -> tuple[bool, str]:
    """POST updated domain settings back to QAD."""
    url = (
        f"{CONFIG['qad']['base_url']}/api/erp/mfgCustomers"
        f"?domainContext={domain}&customerCode={customer_code}"
        f"&viewUri=urn:be:com.qad.base.customer.IMfgCustomer"
    )

    resp = requests.post(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/json",
        },
        json=payload,
        timeout=30,
    )

    if resp.status_code == 401:
        raise _TokenExpired()

    resp_json = resp.json()

    # ── DEBUG BLOCK (remove before production) ────────────────────────────
    print("\n" + "="*60)
    print("DEBUG: POST mfgCustomers response")
    print("="*60)
    print(f"  HTTP Status   : {resp.status_code}")
    print(f"  Customer      : {customer_code}")
    print(f"  Domain        : {domain}")
    print(f"  Submit Success: {resp_json.get('submitResult', {}).get('success')}")
    errors = resp_json.get("submitResult", {}).get("errors", [])
    print(f"  Errors ({len(errors)}):")
    for e in errors:
        print(f"    - field   : {e.get('field', 'N/A')}")
        print(f"      message : {e.get('message', 'N/A')}")
        print(f"      value   : {e.get('value', 'N/A')}")
        print(f"      code    : {e.get('code', 'N/A')}")
    print("\n  Full submitResult:")
    print(json.dumps(resp_json.get("submitResult", {}), indent=4))
    print("="*60 + "\n")
    # ── END DEBUG BLOCK ───────────────────────────────────────────────────

    submit = resp_json.get("submitResult", {})

    if submit.get("success") is True:
        return True, ""

    errors    = submit.get("errors", [])
    error_msg = "; ".join(
        e.get("message", "").strip()
        for e in errors
        if e.get("message", "").strip()
    ) or f"HTTP {resp.status_code} — submitResult.success was not True"

    return False, error_msg


def update_domain_settings(
    customer_code: str,
    domain:        str,
    site_code:     str,
    daybook_set:   str,
    token:         str,
) -> tuple[bool, str]:
    """
    GET auto-created mfgCustomer → patch siteCode + daybookSetCode → POST back.
    Returns (success, error_msg).
    """
    existing = get_mfg_customer(domain, customer_code, token)

    mfg_list = (
        existing.get("data", {}).get("mfgCustomers")
        or existing.get("mfgCustomers")
    )

    if not mfg_list:
        return False, f"Domain settings record not found for customer '{customer_code}' in domain '{domain}'"

    payload    = existing.get("data", existing)
    mfg_record = mfg_list[0]

    mfg_record["siteCode"]       = site_code
    mfg_record["daybookSetCode"] = daybook_set

    return post_mfg_customer(payload, domain, customer_code, token)


# =============================================================================
# 5. WORKBOOK HELPERS
# =============================================================================

def _mark_done(ws, row_idx: int, status_col: int, error_col: int):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col,  value="")


def _mark_error(
    ws,
    row_idx:    int,
    status_col: int,
    error_col:  int,
    header_row: list,
    bad_cols:   list,
    error_msg:  str,
):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL

    ws.cell(row=row_idx, column=1).fill          = RED_FILL
    ws.cell(row=row_idx, column=status_col).fill = RED_FILL
    ws.cell(row=row_idx, column=error_col).fill  = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="ERROR")
    ws.cell(row=row_idx, column=error_col,  value=error_msg)

    for col_name in bad_cols:
        if col_name in header_row:
            ws.cell(row=row_idx, column=header_row.index(col_name) + 1).fill = RED_FILL


# =============================================================================
# 6. MANDATORY FIELD CHECK
# =============================================================================

def _check_mandatory(row_data: dict) -> list[str]:
    missing = []
    for col in MANDATORY_COLUMNS:
        v = row_data.get(col)
        if v is None or str(v).strip() == "" or str(v).strip().lower() == "none":
            missing.append(col)
    return missing


def _check_mandatory_domain(row_data: dict) -> list[str]:
    missing = []
    for col in MANDATORY_DOMAIN_COLUMNS:
        v = row_data.get(col)
        if v is None or str(v).strip() == "" or str(v).strip().lower() == "none":
            missing.append(col)
    return missing


# =============================================================================
# 7. SINGLE-FILE PROCESSOR
# =============================================================================

def process_file(file_path: str, tm: TokenManager) -> tuple[int, int]:
    """Open one workbook, process every row, return (ok_count, fail_count)."""

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Row 1 = section labels, Row 2 = column headers, Row 3+ = data
    header_row = [
        str(c.value).strip() if c.value is not None else ""
        for c in ws[2]
    ]

    for col_name in ("Status", "Error"):
        if col_name not in header_row:
            ws.cell(row=2, column=len(header_row) + 1, value=col_name)
            header_row.append(col_name)

    status_col = header_row.index("Status") + 1
    error_col  = header_row.index("Error")  + 1

    ok_count   = 0
    fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3), start=3):

        row_values = [cell.value for cell in row]

        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        status         = str(row_data.get("Status", "")).strip().upper()
        data_operation = str(row_data.get("Data Operation", "")).strip().upper() or "C"

        if status == "DONE" and data_operation == "C":
            continue

        # ── Mandatory check ───────────────────────────────────────────────
        missing = _check_mandatory(row_data)
        if missing:
            fail_count += 1
            _mark_error(
                ws, row_idx, status_col, error_col, header_row,
                missing,
                f"Missing mandatory fields: {', '.join(missing)}",
            )
            wb.save(file_path)
            continue

        # ── Operation validity check ──────────────────────────────────────
        if data_operation not in {"C", "U"}:
            fail_count += 1
            _mark_error(
                ws, row_idx, status_col, error_col, header_row,
                ["Data Operation"],
                "Data Operation must be C, U, or blank",
            )
            wb.save(file_path)
            continue

        # ── Domain settings mandatory check (CREATE only) ─────────────────
        is_create = (data_operation == "C")

        if is_create:
            missing_domain = _check_mandatory_domain(row_data)
            if missing_domain:
                fail_count += 1
                _mark_error(
                    ws, row_idx, status_col, error_col, header_row,
                    missing_domain,
                    f"Missing domain setting fields: {', '.join(missing_domain)}",
                )
                wb.save(file_path)
                continue

        # ── Build payload ─────────────────────────────────────────────────
        if data_operation == "U":
            existing = get_customer(
                row_data.get("Shared Set", ""),
                row_data.get("Customer", ""),
                tm.get(),
            )

            if not existing.get("data") or not existing["data"].get("customerV2s"):
                fail_count += 1
                _mark_error(
                    ws, row_idx, status_col, error_col, header_row, [],
                    f"UPDATE failed: Customer '{row_data.get('Customer', '')}' not found in QAD",
                )
                wb.save(file_path)
                continue

            payload  = existing["data"]
            customer = payload["customerV2s"][0]

            customer["businessRelationCode"]           = str(row_data.get("Business Relation", "")).strip()
            customer["isActive"]                       = str(row_data.get("Active", "")).strip().lower() == "yes"
            customer["currencyCode"]                   = str(row_data.get("Currency", "")).strip()
            customer["creditTermsCode"]                = str(row_data.get("Credit Terms", "")).strip()
            customer["creditRatingCode"]               = str(row_data.get("Credit Rating")).strip()
            customer["invoiceStatusCode"]              = str(row_data.get("Invoice Status", "")).strip()
            customer["taxZone"]                        = str(row_data.get("Tax Zone", "")).strip()
            customer["telephone"]                      = str(row_data.get("Telephone")).strip()
            customer["EMail"]                            = str(row_data.get("Email")).strip()
            customer["invoiceControlGLProfileCode"]    = str(row_data.get("Invoice Control GL Profile", "")).strip()
            customer["creditNoteControlGLProfileCode"] = str(row_data.get("Credit Note Control GL Profile", "")).strip()
            customer["prePaymentControlGLProfileCode"] = str(row_data.get("Prepayment Control GL Profile", "")).strip()
            customer["salesAccountGLProfileCode"]      = str(row_data.get("Sales Account GL Profile", "")).strip()

        else:
            payload = build_payload(row_data)

        # ── Step 1: POST customer (with one token-refresh retry) ──────────
        success   = False
        error_msg = ""

        for attempt in range(2):
            try:
                success, error_msg = post_customer(payload, tm.get(), is_create=is_create)
                break
            except _TokenExpired:
                if attempt == 0:
                    tm.refresh()
                    continue
                error_msg = "Token refresh failed — unauthorised"
                break
            except requests.RequestException as e:
                error_msg = f"Network error: {e}"
                break

        if not success:
            fail_count += 1
            _mark_error(ws, row_idx, status_col, error_col, header_row, [], error_msg)
            wb.save(file_path)
            time.sleep(0.1)
            continue

        # ── Step 2 (CREATE only): update domain settings ──────────────────
        if is_create:
            domain        = str(row_data.get("Domain", "")).strip()
            site_code     = str(row_data.get("Site Code", "")).strip()
            daybook_set   = str(row_data.get("Daybook Set", "")).strip()
            customer_code = str(row_data.get("Customer", "")).strip()

            domain_ok  = False
            domain_err = ""

            for attempt in range(2):
                try:
                    domain_ok, domain_err = update_domain_settings(
                        customer_code, domain, site_code, daybook_set, tm.get()
                    )
                    break
                except _TokenExpired:
                    if attempt == 0:
                        tm.refresh()
                        continue
                    domain_err = "Token refresh failed — unauthorised"
                    break
                except requests.RequestException as e:
                    domain_err = f"Network error: {e}"
                    break

            if not domain_ok:
                fail_count += 1
                _mark_error(
                    ws, row_idx, status_col, error_col, header_row, [],
                    f"Customer created but domain settings failed: {domain_err}",
                )
                wb.save(file_path)
                time.sleep(0.1)
                continue

        # ── All steps passed ──────────────────────────────────────────────
        ok_count += 1
        _mark_done(ws, row_idx, status_col, error_col)
        wb.save(file_path)
        time.sleep(0.1)

    return ok_count, fail_count


# =============================================================================
# 8. ORCHESTRATOR
# =============================================================================

def run(folder_path: str) -> tuple[int, int]:
    folder = os.path.abspath(folder_path)

    if not os.path.exists(folder):
        print(f"ERROR: Folder not found: {folder}")
        raise RuntimeError(f"Folder not found: {folder}")

    xlsx_files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not xlsx_files:
        print(f"ERROR: No .xlsx files found in: {folder}")
        raise RuntimeError(f"No .xlsx files found in: {folder}")

    tm = TokenManager()

    total_ok   = 0
    total_fail = 0

    for file_name in xlsx_files:
        file_path = os.path.join(folder, file_name)
        print(f"Loading : {file_path}")

        ok, fail   = process_file(file_path, tm)
        total_ok   += ok
        total_fail += fail

        print("\nLoad Summary")
        print("-" * 40)
        print(f"  Rows loaded successfully : {ok}")
        print(f"  Rows failed              : {fail}")
        if fail == 0:
            print("  Result : ALL ROWS LOADED SUCCESSFULLY ✓")
        else:
            print("  Result : COMPLETED WITH ERRORS — fix red rows and re-run")
        print()

    print("=" * 40)
    print(f"  Total loaded : {total_ok}")
    print(f"  Total failed : {total_fail}")
    print("=" * 40)

    return total_ok, total_fail


# =============================================================================
# 9. ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    folder = os.path.abspath(
        os.path.join(ROOT_DIR, CONFIG["folders"]["customer"])
    )
    ok, fail = run(folder)
    sys.exit(0 if fail == 0 else 1)