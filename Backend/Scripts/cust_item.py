import openpyxl
import requests
import time
import sys
import os
from openpyxl.styles import PatternFill

# ==========================================
# AUTH SETUP
# ==========================================
TOKEN_URL  = "https://cat5-devl.adaptive.qad.com/clouderp/oauth/token"
UPLOAD_URL = "https://cat5-devl.adaptive.qad.com/clouderp/api/erp/customerItemV2s?viewUri=urn:be:com.qad.sales.item.ICustomerItemV2"

AUTH_PARAMS = {
    "client_id": "afb97fd221925b87f01489aeb0e02e81",
    "username": "demo",
    "password": "qad",
    "grant_type": "password"
}

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)


def get_new_token():
    try:
        response = requests.post(TOKEN_URL, params=AUTH_PARAMS)
        response.raise_for_status()
        token = response.json().get("access_token")
        if token:
            return token
        print("❌ Failed to obtain access token.")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Token error: {e}")
        sys.exit(1)


def ensure_columns(ws):
    header_row = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
    for name in ("Status", "Error"):
        if name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=name)
            header_row.append(name)
    return header_row, header_row.index("Status") + 1, header_row.index("Error") + 1


def mark_error(ws, row_idx, status_col, error_col, error_msg):
    ws.cell(row=row_idx, column=1).fill = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="")
    ws.cell(row=row_idx, column=error_col, value=error_msg)
    ws.cell(row=row_idx, column=error_col).fill = RED_FILL


def mark_success(ws, row_idx, status_col, error_col):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col, value="")


def rename_error(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if not name.startswith("error_"):
        new_path = os.path.join(folder, "error_" + name)
        os.rename(file_path, new_path)
        return new_path
    return file_path


def rename_restore(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if name.startswith("error_"):
        new_path = os.path.join(folder, name[len("error_"):])
        os.rename(file_path, new_path)
        return new_path
    return file_path


def run(file_path):
    print(f"\n{'─'*55}")
    print(f"  📄 {os.path.basename(file_path)}")
    print(f"{'─'*55}")

    current_token = get_new_token()
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    header_row, status_col, error_col = ensure_columns(ws)

    success_count = skip_count = fail_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        row_data = dict(zip(header_row, row_values))

        if str(row_data.get("Status", "")).strip().upper() == "DONE":
            ci = str(row_data.get("Customer Item Code", "")).strip()
            print(f"  ⏭  {ci or f'row {row_idx}'} — already DONE")
            skip_count += 1
            continue

        domain_code        = str(row_data.get("Domain Code", "")).strip()
        item_code          = str(row_data.get("Item Code", "")).strip()
        customer_item_code = str(row_data.get("Customer Item Code", "")).strip()

        missing = []
        if not item_code:          missing.append("Item Code")
        if not customer_item_code: missing.append("Customer Item Code")

        if missing:
            msg = f"Missing required field(s): {', '.join(missing)}"
            mark_error(ws, row_idx, status_col, error_col, msg)
            print(f"  ✘  row {row_idx} — {msg}")
            fail_count += 1
            continue

        uri = f"urn:be:com.qad.sales.item.ICustomerItemV2:{domain_code}.."

        payload = {
            "supplementaryMessages": [],
            "customerItemV2s": [{
                "uri":                    uri,
                "concurrencyHash":        "",
                "cpMstrUser1":            "",
                "cpMstrUser2":            "",
                "creatingUser":           "",
                "customDecimal0": 0, "customDecimal1": 0, "customDecimal2": 0,
                "customDecimal3": 0, "customDecimal4": 0,
                "customInteger0": 0, "customInteger1": 0, "customInteger2": 0,
                "customInteger3": 0, "customInteger4": 0,
                "customLong0": "", "customLong1": "",
                "customNote": "",
                "customShort0":  "", "customShort1":  "", "customShort2":  "",
                "customShort3":  "", "customShort4":  "", "customShort5":  "",
                "customShort6":  "", "customShort7":  "", "customShort8":  "",
                "customShort9":  "", "customShort10": "", "customShort11": "",
                "customShort12": "", "customShort13": "", "customShort14": "",
                "customShort15": "", "customShort16": "", "customShort17": "",
                "customShort18": "", "customShort19": "",
                "customerAddressField1": "", "customerAddressField2": "",
                "customerAddressField3": "", "customerAddressField4": "",
                "customerAddressField5": "", "customerAddressField6": "",
                "customerCode":           "",
                "customerItemCode":       customer_item_code,
                "customerItemDescription": "",
                "customerItemECONbr":     "",
                "dataOperation":          "",
                "displayCustomerItem":    "",
                "domainCode":             domain_code,
                "itemCode":               item_code,
                "itemDescription1":       "",
                "itemDescription2":       "",
                "lastModifiedUser":       "",
                "negativePlanVariance":   0,
                "negativeShipVariance":   0,
                "packageCode":            "",
                "positivePlanVariance":   0,
                "positiveShipVariance":   0
            }]
        }

        retry = True
        while retry:
            try:
                response = requests.post(
                    UPLOAD_URL, json=payload,
                    headers={"Content-Type": "application/json", "Authorization": f"Bearer {current_token}"}
                )

                if response.status_code == 401:
                    current_token = get_new_token()
                    continue

                resp_json = response.json()

                if response.status_code == 200 and resp_json.get("submitResult", {}).get("success"):
                    mark_success(ws, row_idx, status_col, error_col)
                    print(f"  ✔  Customer Item: {customer_item_code} / Item: {item_code}")
                    success_count += 1
                else:
                    errors = resp_json.get("submitResult", {}).get("errors", [])
                    msgs = []
                    for e in errors:
                        msg = e.get("message", "")
                        if "already exists" in msg.lower():
                            msgs.append(f"Duplicate — Customer Item '{customer_item_code}' already exists")
                        elif e.get("fieldName"):
                            msgs.append(f"Field '{e['fieldName']}': {msg}")
                        else:
                            msgs.append(msg)
                    error_msg = "; ".join(msgs) or resp_json.get("message", f"HTTP {response.status_code}")
                    mark_error(ws, row_idx, status_col, error_col, error_msg)
                    print(f"  ✘  Customer Item: {customer_item_code} — {error_msg}")
                    fail_count += 1

                retry = False

            except Exception as e:
                mark_error(ws, row_idx, status_col, error_col, str(e))
                print(f"  ✘  Customer Item: {customer_item_code} — Connection error: {e}")
                fail_count += 1
                retry = False

        time.sleep(0.1)

    wb.save(file_path)

    if fail_count > 0:
        file_path = rename_error(file_path)
        print(f"\n  ⚠  Errors found — file renamed to: {os.path.basename(file_path)}")
    elif success_count > 0 and fail_count == 0:
        file_path = rename_restore(file_path)

    print(f"\n  📊 Success: {success_count} | Skipped: {skip_count} | Failed: {fail_count}")
    return success_count, fail_count


# ==========================================
# ENTRY POINT
# ==========================================
if __name__ == "__main__":
    base   = os.path.dirname(os.path.abspath(__file__))
    folder = os.path.join(base, "..", "Customer_Item")

    if not os.path.exists(folder):
        print(f"❌ Folder not found: {folder}")
        sys.exit(1)

    files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not files:
        print("⚠  No .xlsx files found in Customer_Item/")
        sys.exit(0)

    total_s = total_f = 0
    for f in files:
        s, fail = run(os.path.join(folder, f))
        total_s += s
        total_f += fail

    print(f"\n{'═'*55}")
    print(f"  TOTAL — Success: {total_s} | Failed: {total_f}")
    print(f"{'═'*55}")
    sys.exit(0 if total_f == 0 else 1)