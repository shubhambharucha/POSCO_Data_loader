import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL = PatternFill(fill_type=None)

HEADER_RULES = {
    "Domain Code":           {"type": "character", "max_len": 8},
    "SO Number":             {"type": "character", "max_len": 8},
    "Sold To Customer Code": {"type": "character", "max_len": 8},
    "Bill To Customer Code": {"type": "character", "max_len": 8},
    "Ship To Customer Code": {"type": "character", "max_len": 8},
    "Site Code":             {"type": "character", "max_len": 8},
    "Currency":              {"type": "character", "max_len": 3},
    "Daybook Set":           {"type": "character", "max_len": 8},
    "Credit Terms":          {"type": "character", "max_len": 8},
    "Order Date":            {"type": "date"},
    "Due Date":              {"type": "date"},
    "Ship Via":              {"type": "character", "max_len": 20},
    "Freight List":          {"type": "character", "max_len": 8, "optional": True},
    "Freight Terms":         {"type": "character", "max_len": 20, "optional": True},
}

LINES_RULES = {
    "SO Number":        {"type": "character", "max_len": 8},
    "Item Code":        {"type": "character", "max_len": None},
    "Site Code":        {"type": "character", "max_len": 8},
    "Quantity Ordered": {"type": "decimal",   "min": 0},
    "List Price":       {"type": "decimal",   "min": 0},
    "Discount":         {"type": "decimal",   "min": 0,   "optional": True},
    "Net Price":        {"type": "decimal",   "min": 0,   "optional": True},
    "Due Date":         {"type": "date"},
}

HEADER_ENTITY_COL = "SO Number"
LINES_ENTITY_COL  = "SO Number"
SKIP_STATUSES     = {"DONE", "READY"}


def _check_char(value, max_len):
    if value is None or str(value).strip() == "" or str(value).strip().lower() == "none":
        return "empty"
    sv = str(value).strip()
    if max_len and len(sv) > max_len:
        return f"max {max_len} chars (got {len(sv)})"
    return None


def _check_decimal(value, min_val=0):
    if value is None or str(value).strip() == "":
        return "empty"
    try:
        if float(value) < min_val:
            return f"must be >= {min_val}"
    except (ValueError, TypeError):
        return "invalid number"
    return None


def _check_integer(value):
    if value is None or str(value).strip() == "":
        return "empty"
    try:
        int(float(value))
    except (ValueError, TypeError):
        return "invalid integer"
    return None


def _check_date(value):
    if value is None or str(value).strip() == "":
        return "empty"
    if isinstance(value, datetime):
        return None
    try:
        float(value)  # Excel serial date
        return None
    except (ValueError, TypeError):
        pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S.000Z"):
        try:
            datetime.strptime(str(value).strip(), fmt)
            return None
        except ValueError:
            continue
    return "invalid date"

# ----- def validate_freight_dependency(row_data, row_header):
# Add the future when you want to scale

def _validate_sheet(ws, rules, entity_col):
    """Validates a single sheet against the given rules. Returns per-sheet stats dict."""

    raw_headers = [cell.value for cell in ws[1]]
    header_row  = [str(h).strip() if h is not None else "" for h in raw_headers]

    if "Status" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Status")
        header_row.append("Status")
    if "Error" not in header_row:
        ws.cell(row=1, column=len(header_row) + 1, value="Error")
        header_row.append("Error")

    status_col_idx = header_row.index("Status") + 1
    error_col_idx  = header_row.index("Error")  + 1
    entity_col_idx = header_row.index(entity_col) + 1 if entity_col in header_row else None

    rows_processed = 0
    rows_passed    = 0
    rows_failed    = 0
    rows_skipped   = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]

        if not any(row_values):
            rows_skipped += 1
            continue

        row_data = dict(zip(header_row, row_values))

        current_status = str(row_data.get("Status", "") or "").strip().upper()
        if current_status in SKIP_STATUSES:
            rows_skipped += 1
            continue

        rows_processed += 1
        row_errors      = []
        error_cell_idxs = []

        for col_name, rule in rules.items():
            if col_name not in header_row:
                continue

            col_idx  = header_row.index(col_name) + 1
            value    = row_data.get(col_name)
            optional = rule.get("optional", False)

            # Optional fields: skip entirely if empty
            if optional:
                if value is None or str(value).strip() == "":
                    continue

            if rule["type"] == "character":
                err = _check_char(value, rule.get("max_len"))
            elif rule["type"] == "decimal":
                err = _check_decimal(value, rule.get("min", 0))
            elif rule["type"] == "integer":
                err = _check_integer(value)
            elif rule["type"] == "date":
                err = _check_date(value)
            else:
                err = None

            if err:
                row_errors.append(f"{col_name}: {err}")
                error_cell_idxs.append(col_idx)

# ----------------------------------------------------------------------------------------------
# Freight Dependency Validation 
# -------------------------------------------------------
        if "Freight List" in header_row and "Freight Terms" in header_row:

            freight_list = str(row_data.get("Freight List", "") or "").strip()
            freight_terms = str(row_data.get("Freight Terms", "") or "").strip()

            if bool(freight_list) != bool(freight_terms):
                row_errors.append("Freight List and Freight Terms must both be filled or both be empty")
                error_cell_idxs.append(header_row.index("Freight List") + 1)
                error_cell_idxs.append(header_row.index("Freight Terms") + 1)

#--------------------------------------------------------------------------------------------
#Clear old formatting, and writing the results
#--------------------------------------------------------------------------------------------

        for cell in ws[row_idx]:
            cell.fill = CLEAR_FILL

        if row_errors:
            rows_failed += 1
            error_msg   = "; ".join(row_errors)

            if entity_col_idx:
                ws.cell(row=row_idx, column=entity_col_idx).fill = RED_FILL
            for cidx in error_cell_idxs:
                ws.cell(row=row_idx, column=cidx).fill = RED_FILL
            ws.cell(row=row_idx, column=error_col_idx,  value=error_msg).fill = RED_FILL
            ws.cell(row=row_idx, column=status_col_idx, value="ERROR")
        else:
            rows_passed += 1
            ws.cell(row=row_idx, column=status_col_idx, value="READY")
            ws.cell(row=row_idx, column=error_col_idx,  value="")

    return {
        "rows_processed": rows_processed,
        "rows_passed":    rows_passed,
        "rows_failed":    rows_failed,
        "rows_skipped":   rows_skipped,
    }


def validate(file_path):
    wb = openpyxl.load_workbook(file_path)

    totals = {
        "has_errors":     False,
        "rows_processed": 0,
        "rows_passed":    0,
        "rows_failed":    0,
        "rows_skipped":   0,
    }

    if "Header" in wb.sheetnames:
        stats = _validate_sheet(wb["Header"], HEADER_RULES, HEADER_ENTITY_COL)
        for key in ("rows_processed", "rows_passed", "rows_failed", "rows_skipped"):
            totals[key] += stats[key]

    if "Lines" in wb.sheetnames:
        stats = _validate_sheet(wb["Lines"], LINES_RULES, LINES_ENTITY_COL)
        for key in ("rows_processed", "rows_passed", "rows_failed", "rows_skipped"):
            totals[key] += stats[key]

    if totals["rows_failed"] > 0:
        totals["has_errors"] = True

    wb.save(file_path)
    return totals


# ---------------------------------------------------------------------------
# Standalone testing mode
# ---------------------------------------------------------------------------
def run_standalone():

    import sys
    import os

    #Temporary fix for standalone execution in cmd
    ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    
    sys.path.append(ROOT_DIR)

    from config import CONFIG

    folder = os.path.abspath(os.path.join(ROOT_DIR, CONFIG["folders"]["sales_order"]))

    if not os.path.exists(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)
    
    xlsx_files = [f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")]

    if not xlsx_files:
        print(f"ERROR: No .xlsx file found in folder: {folder}")
        sys.exit(1)

    total_processed = 0
    total_failed   = 0

    for file_name in xlsx_files:

        file_path = os.path.join(folder, file_name)

        print(f"Validating: {file_path} ...")

        result = validate(file_path)

        print("\nValidation Summary")
        print("------------------------------")
        print(f"Rows processed : {result['rows_processed']}")
        print(f"Rows passed    : {result['rows_passed']}")
        print(f"Rows failed    : {result['rows_failed']}")
        print(f"Rows skipped   : {result['rows_skipped']}")

        total_processed += result["rows_processed"]
        total_failed    += result["rows_failed"]

        if result["has_errors"]:
            print("\nValidation FAILED")
        else:
            print("\nValidation PASSED — all rows are READY for Loading.")

    #outside loop    
    sys.exit(0 if total_failed == 0 else 1)

if __name__ == "__main__":
    run_standalone()