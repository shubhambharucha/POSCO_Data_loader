import openpyxl
import requests
import time
import sys
import os
from openpyxl.styles import PatternFill
from datetime import datetime
from collections import defaultdict

# ==========================================
# AUTH SETUP
# ==========================================
BASE_URL   = "https://cat5-devl.adaptive.qad.com/clouderp"
TOKEN_URL  = f"{BASE_URL}/oauth/token"

AUTH_PARAMS = {
    "client_id": "afb97fd221925b87f01489aeb0e02e81",
    "username":  "demo",
    "password":  "qad",
    "grant_type":"password"
}

HEADER_URL       = f"{BASE_URL}/api/erp/purchaseOrderHeaders?viewUri=urn:be:com.qad.purchasing.purchaseorders.IPurchaseOrderHeader"
INIT_LINE_URL    = f"{BASE_URL}/api/erp/purchaseOrderLinesGrid?initialize=true&domainCode={{domain}}&purchaseOrderNumber={{po}}"
FIELD_CHANGE_URL = f"{BASE_URL}/api/erp/purchaseOrderLines/fieldChangeV2?fieldName={{fieldName}}"
IS_RECEIVED_URL  = f"{BASE_URL}/api/erp/purchaseOrderLines/isReceivedPurchaseOrderLineV2?domainCode={{domain}}&purchaseOrderNumber={{po}}&purchaseOrderLine={{line}}"
SYNC_LINE_URL    = f"{BASE_URL}/api/erp/purchaseOrderLinesGrid"

RED_FILL     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CLEAR_FILL   = PatternFill(fill_type=None)
DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]


# ==========================================
# HELPERS
# ==========================================

def get_new_token():
    try:
        response = requests.post(TOKEN_URL, params=AUTH_PARAMS)
        response.raise_for_status()
        token = response.json().get("access_token")
        if token:
            return token
        print("❌ Failed to obtain access token.")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Token error: {e}")
        sys.exit(1)


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT00:00:00.000Z")
    str_val = str(val).strip()
    if not str_val or str_val.lower() == "none":
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(str_val, fmt).strftime("%Y-%m-%dT00:00:00.000Z")
        except ValueError:
            continue
    return str_val


def sv(row_data, key, default=""):
    val = row_data.get(key, default)
    return str(val).strip() if val is not None else default


def fv(row_data, key, default=0.0):
    val = row_data.get(key, default)
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def iv(row_data, key, default=0):
    val = row_data.get(key, default)
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return default


def to_iso_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%dT%H:%M:%S.000Z')
    elif val and isinstance(val, str):
        try:
            dt = datetime.strptime(val, '%Y-%m-%d %H:%M:%S')
            return dt.strftime('%Y-%m-%dT%H:%M:%S.000Z')
        except ValueError:
            return val
    return val


def make_headers(token):
    return {
        "Content-Type":  "application/json",
        "Authorization": f"Bearer {token}"
    }


def get_with_retry(url, token_ref):
    while True:
        resp = requests.get(url, headers=make_headers(token_ref[0]))
        if resp.status_code == 401:
            token_ref[0] = get_new_token()
            continue
        try:
            return resp, resp.json()
        except Exception:
            return resp, {}


def post_with_retry(url, payload, token_ref):
    while True:
        resp = requests.post(url, json=payload, headers=make_headers(token_ref[0]))
        if resp.status_code == 401:
            token_ref[0] = get_new_token()
            continue
        try:
            return resp, resp.json()
        except Exception:
            return resp, {}


def mark_row_error(ws, row_idx, status_col, error_col, error_msg):
    """Red fill col 1 and Error cell only. Clear other fills."""
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=1).fill = RED_FILL
    ws.cell(row=row_idx, column=status_col, value="ERROR")
    ws.cell(row=row_idx, column=error_col, value=error_msg)
    ws.cell(row=row_idx, column=error_col).fill = RED_FILL


def mark_row_success(ws, row_idx, status_col, error_col):
    for cell in ws[row_idx]:
        cell.fill = CLEAR_FILL
    ws.cell(row=row_idx, column=status_col, value="DONE")
    ws.cell(row=row_idx, column=error_col, value="")


def ensure_columns(ws, *col_names):
    header_row = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in ws[1]
    ]
    for name in col_names:
        if name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=name)
            header_row.append(name)
    return header_row


def rename_error(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if not name.startswith("error_"):
        new_path = os.path.join(folder, "error_" + name)
        os.rename(file_path, new_path)
        return new_path
    return file_path


def rename_restore(file_path):
    folder = os.path.dirname(file_path)
    name   = os.path.basename(file_path)
    if name.startswith("error_"):
        new_path = os.path.join(folder, name[len("error_"):])
        os.rename(file_path, new_path)
        return new_path
    return file_path


def parse_api_errors(resp_json, entity_name=""):
    errors = resp_json.get("submitResult", {}).get("errors", [])
    msgs = []
    for e in errors:
        msg = e.get("message", "")
        if "already exists" in msg.lower():
            msgs.append(f"Duplicate — '{entity_name}' already exists")
        elif e.get("fieldName"):
            msgs.append(f"Field '{e['fieldName']}': {msg}")
        else:
            msgs.append(msg)
    return "; ".join(msgs) or resp_json.get("message", "Unknown error")


# ==========================================
# PAYLOAD BUILDERS
# ==========================================

def build_header_payload(row_data):
    domain   = sv(row_data, "Domain Code")
    po_num   = sv(row_data, "PO Number")
    order_dt = to_iso_date(row_data.get("Order Date")) or to_iso_date(datetime.now())
    due_dt   = to_iso_date(row_data.get("Due Date"))   or to_iso_date(datetime.now())

    return {
        "purchaseOrderHeaders": [{
            "purchaseOrderNumber":    po_num,
            "domainCode":             domain,
            "supplierCode":           sv(row_data, "Supplier Code"),
            "currencyCode":           sv(row_data, "Currency"),
            "exchangeRate":           1,
            "exchangeRate2":          1,
            "exchangeRateType":       "ACCOUNTING",
            "creditTermsCode":        sv(row_data, "Credit Terms"),
            "daybookSetCode":         sv(row_data, "Daybook Set"),
            "shipToSite":             sv(row_data, "Ship To Site"),
            "billToCode":             sv(row_data, "Bill To Code"),
            "taxEnvironment":         sv(row_data, "Tax Environment", "USA"),
            "languageCode":           sv(row_data, "Language Code", "us"),
            "contact":                sv(row_data, "Contact"),
            "remarks":                sv(row_data, "Remarks"),
            "orderDate":              order_dt,
            "dueDate":                due_dt,
            "pricingDate":            order_dt,
            "lastPriceDate":          order_dt,
            "startEffective":         order_dt,
            "endEffective":           order_dt,
            "orderRevisionDate":      order_dt,
            "taxDateSummary":         order_dt,
            "orderStatus":            "O",
            "isConfirmed":            True,
            "isFixedPrice":           True,
            "isPrintPurchaseOrder":   True,
            "isTaxable":              True,
            "isDisplayTaxAmounts":    True,
            "isUsingConsignmentInventory": True,
            "isIntrastatUsed":        True,
            "isPredefaulted":         True,
            "maximumAgingDays":       90,
            "transmitFlag":           "1",
            "ERSOption":              "1",
            "ERSPriceListOption":     "0",
            "dataOperation":          "C",
            "concurrencyHash":        "",
            "disallowedActions":      "",
            "disallowedActionsMessage": "",
            "supplementaryMessages":  [],
            "POIntrastats":           [],
        }]
    }


# ==========================================
# LINE CREATION FLOW
# ==========================================

def create_line(domain, po_num, line_row_data, line_number, token_ref):
    site_code = sv(line_row_data, "Site Code")
    item_code = sv(line_row_data, "Item Code")
    qty       = fv(line_row_data, "Quantity Ordered")
    price     = fv(line_row_data, "Unit Price")
    due_dt    = to_iso_date(line_row_data.get("Due Date")) or to_iso_date(datetime.now())

    # 1. Initialize blank line
    print(f"    ↳ Init line {line_number}...")
    init_url = INIT_LINE_URL.format(domain=domain, po=po_num)
    resp, resp_json = get_with_retry(init_url, token_ref)

    if resp.status_code != 200:
        return False, f"Init failed: HTTP {resp.status_code}"

    lines = resp_json.get("data", {}).get("purchaseOrderLines", [])
    if not lines:
        return False, "Init returned no line object"

    line = lines[0]
    line["purchaseOrderLine"] = line_number
    line["dueDate"]           = due_dt

    # 2. fieldChange: siteCode
    line["siteCode"] = site_code
    resp, resp_json = post_with_retry(
        FIELD_CHANGE_URL.format(fieldName="siteCode"),
        {"purchaseOrderLines": [line]}, token_ref
    )
    if resp.status_code != 200:
        return False, f"fieldChange(siteCode) failed: HTTP {resp.status_code}"
    lines = resp_json.get("data", {}).get("purchaseOrderLines", [])
    if not lines:
        return False, "fieldChange(siteCode) returned no line"
    line = lines[0]

    # 3. fieldChange: itemCode
    line["itemCode"] = item_code
    resp, resp_json = post_with_retry(
        FIELD_CHANGE_URL.format(fieldName="itemCode"),
        {"purchaseOrderLines": [line]}, token_ref
    )
    if resp.status_code != 200:
        return False, f"fieldChange(itemCode) failed: HTTP {resp.status_code}"
    lines = resp_json.get("data", {}).get("purchaseOrderLines", [])
    if not lines:
        return False, "fieldChange(itemCode) returned no line"
    line = lines[0]

    time.sleep(0.1)

    # 4. isReceived check
    is_recv_url = IS_RECEIVED_URL.format(domain=domain, po=po_num, line=line_number)
    get_with_retry(is_recv_url, token_ref)

    time.sleep(0.1)

    # 5. fieldChange: quantityOrdered
    line["quantityOrdered"] = qty
    resp, resp_json = post_with_retry(
        FIELD_CHANGE_URL.format(fieldName="quantityOrdered"),
        {"purchaseOrderLines": [line]}, token_ref
    )
    if resp.status_code != 200:
        return False, f"fieldChange(quantityOrdered) failed: HTTP {resp.status_code}"
    lines = resp_json.get("data", {}).get("purchaseOrderLines", [])
    if not lines:
        return False, "fieldChange(quantityOrdered) returned no line"
    line = lines[0]

    # 6. fieldChange: purchaseCost
    line["purchaseCost"] = price
    resp, resp_json = post_with_retry(
        FIELD_CHANGE_URL.format(fieldName="purchaseCost"),
        {"purchaseOrderLines": [line]}, token_ref
    )
    if resp.status_code != 200:
        return False, f"fieldChange(purchaseCost) failed: HTTP {resp.status_code}"
    lines = resp_json.get("data", {}).get("purchaseOrderLines", [])
    if not lines:
        return False, "fieldChange(purchaseCost) returned no line"
    line = lines[0]

    # 7. Sync / commit line
    resp, resp_json = post_with_retry(
        SYNC_LINE_URL,
        {"purchaseOrderLines": [line]}, token_ref
    )
    if resp.status_code == 200 and resp_json.get("submitResult", {}).get("success"):
        return True, ""
    else:
        error_msg = parse_api_errors(resp_json, f"PO {po_num} Line {line_number}")
        return False, f"Sync failed: {error_msg}"


# ==========================================
# MAIN RUN
# ==========================================

def run(file_path):
    print(f"\n{'─'*55}")
    print(f"  📄 {os.path.basename(file_path)}")
    print(f"{'─'*55}")

    token_ref = [get_new_token()]

    wb = openpyxl.load_workbook(file_path)

    if "Header" not in wb.sheetnames or "Lines" not in wb.sheetnames:
        print("❌ Workbook must have sheets named 'Header' and 'Lines'")
        return 0, 1

    ws_h = wb["Header"]
    ws_l = wb["Lines"]

    h_headers = ensure_columns(ws_h, "Status", "Error")
    l_headers = ensure_columns(ws_l, "Status", "Error")

    h_status_col = h_headers.index("Status") + 1
    h_error_col  = h_headers.index("Error")  + 1
    l_status_col = l_headers.index("Status") + 1
    l_error_col  = l_headers.index("Error")  + 1

    # Index header rows
    header_rows = {}
    for row_idx, row in enumerate(ws_h.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        row_data = dict(zip(h_headers, row_values))
        po_num   = sv(row_data, "PO Number")
        if po_num:
            header_rows[po_num] = (row_idx, row_data)

    # Index line rows
    line_rows = defaultdict(list)
    for row_idx, row in enumerate(ws_l.iter_rows(min_row=2), start=2):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        row_data = dict(zip(l_headers, row_values))
        po_num   = sv(row_data, "PO Number")
        if po_num:
            line_rows[po_num].append((row_idx, row_data))

    h_success = h_skip = h_fail = 0
    l_success = l_skip = l_fail = 0

    for po_num, (h_row_idx, h_row_data) in header_rows.items():

        print(f"\n  📦 PO: {po_num}")

        h_status = str(h_row_data.get("Status", "")).strip().upper()

        # Step 1: Create header
        if h_status == "DONE":
            print(f"  ⏭  Header — already DONE")
            h_skip += 1
        else:
            try:
                payload = build_header_payload(h_row_data)
                resp, resp_json = post_with_retry(HEADER_URL, payload, token_ref)

                if resp.status_code == 200 and resp_json.get("submitResult", {}).get("success"):
                    mark_row_success(ws_h, h_row_idx, h_status_col, h_error_col)
                    h_success += 1
                    print(f"  ✔  Header: {po_num}")
                else:
                    error_msg = parse_api_errors(resp_json, po_num)
                    mark_row_error(ws_h, h_row_idx, h_status_col, h_error_col, error_msg)
                    h_fail += 1
                    print(f"  ✘  Header: {po_num} — {error_msg}")
                    wb.save(file_path)
                    continue

            except Exception as e:
                mark_row_error(ws_h, h_row_idx, h_status_col, h_error_col, str(e))
                h_fail += 1
                print(f"  ✘  Header: {po_num} — {e}")
                wb.save(file_path)
                continue

            time.sleep(0.3)

        # Step 2: Create lines
        domain   = sv(h_row_data, "Domain Code")
        po_lines = line_rows.get(po_num, [])

        if not po_lines:
            print(f"  ⚠  No lines found for PO: {po_num}")
            continue

        for line_number, (l_row_idx, l_row_data) in enumerate(po_lines, start=1):
            l_status = str(l_row_data.get("Status", "")).strip().upper()

            if l_status == "DONE":
                print(f"    ⏭  Line {line_number} — already DONE")
                l_skip += 1
                continue

            explicit_line = iv(l_row_data, "Line Number", 0)
            line_no = explicit_line if explicit_line > 0 else line_number

            item = sv(l_row_data, 'Item Code')
            qty  = fv(l_row_data, 'Quantity Ordered')
            price = fv(l_row_data, 'Unit Price')
            print(f"    → Line {line_no}: {item} | qty={qty} | price={price}")

            try:
                success, error_msg = create_line(domain, po_num, l_row_data, line_no, token_ref)

                if success:
                    mark_row_success(ws_l, l_row_idx, l_status_col, l_error_col)
                    l_success += 1
                    print(f"    ✔  Line {line_no}: {item}")
                else:
                    mark_row_error(ws_l, l_row_idx, l_status_col, l_error_col, error_msg)
                    l_fail += 1
                    print(f"    ✘  Line {line_no}: {item} — {error_msg}")

            except Exception as e:
                mark_row_error(ws_l, l_row_idx, l_status_col, l_error_col, str(e))
                l_fail += 1
                print(f"    ✘  Line {line_no} — {e}")

            wb.save(file_path)
            time.sleep(0.2)

    total_success = h_success + l_success
    total_fail    = h_fail + l_fail

    print(f"\n  📊 Headers — ✔ {h_success} | ⏭ {h_skip} | ✘ {h_fail}")
    print(f"  📊 Lines   — ✔ {l_success} | ⏭ {l_skip} | ✘ {l_fail}")

    if total_fail > 0:
        file_path = rename_error(file_path)
        print(f"\n  ⚠  Errors found — file renamed to: {os.path.basename(file_path)}")
    elif total_success > 0 and total_fail == 0:
        file_path = rename_restore(file_path)

    return total_success, total_fail


# ==========================================
# ENTRY POINT
# ==========================================
if __name__ == "__main__":
    base   = os.path.dirname(os.path.abspath(__file__))
    folder = os.path.join(base, "..", "PurchaseOrder")

    if not os.path.exists(folder):
        print(f"❌ Folder not found: {folder}")
        sys.exit(1)

    files = [
        f for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not files:
        print("⚠  No .xlsx files found in PurchaseOrder/")
        sys.exit(0)

    total_s = total_f = 0
    for f in files:
        s, fail = run(os.path.join(folder, f))
        total_s += s
        total_f += fail

    print(f"\n{'═'*55}")
    print(f"  TOTAL — Success: {total_s} | Failed: {total_f}")
    print(f"{'═'*55}")
    sys.exit(0 if total_f == 0 else 1)